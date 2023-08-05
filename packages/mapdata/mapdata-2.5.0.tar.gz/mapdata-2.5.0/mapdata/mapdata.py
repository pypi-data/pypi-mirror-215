#!/usr/bin/python
#
# mapdata.py
#
# PURPOSE
#	Display a simple interactive map of data points, allowing points to
#	be highlighted by clicking on the map or table or by querying,
#	and allowing some simple data plots.
#
# COPYRIGHT AND LICENSE
#	Copyright (c) 2023, R. Dreas Nielsen
# 	This program is free software: you can redistribute it and/or modify
# 	it under the terms of the GNU General Public License as published by
# 	the Free Software Foundation, either version 3 of the License, or
# 	(at your option) any later version.
# 	This program is distributed in the hope that it will be useful,
# 	but WITHOUT ANY WARRANTY; without even the implied warranty of
# 	MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# 	GNU General Public License for more details.
# 	The GNU General Public License is available at <http://www.gnu.org/licenses/>
#
# AUTHOR
#	Dreas Nielsen (RDN)
#
# ==================================================================

version = "2.5.0"
vdate = "2023-06-20"

copyright = "2023"


import sys
import os.path
import io
import codecs
import argparse
from configparser import ConfigParser
import csv
import re
import datetime
import dateutil.parser as dateparser
import time
import math
import collections
import webbrowser
import threading
import queue
import sqlite3
import tempfile
import subprocess
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tkfont
import tkinter.filedialog as tkfiledialog
import tkintermapview as tkmv
from PIL import ImageGrab
import odf.opendocument
import odf.table
import odf.text
import odf.number
import odf.style
import xlrd
import openpyxl
import numpy as np
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk


# Default name of configuration file.  Files with other names may be read.
config_file_name = "mapdata.conf"

# Configuration files read on startup
config_files = []
# Configuration file read post-startup
config_files_user = []


# Default options
multiselect = "0"
#-- Default location marker.  This may be overridden
location_marker = "triangle_open"
location_color = "black"
use_data_marker = True
use_data_color = True
#-- Selected item marker
select_symbol = "wedge"
select_color = "red"
#-- Label appearance
label_color = "black"
label_font = "Liberation Sans"
label_size = 10
label_bold = False
label_position = "below"	# above or below


# Name of editor, read from environment if it exists
editor = os.environ.get("EDITOR")

#-- Operational configuration
# Whether to use a temporary file for Sqlite (as opposed to memory).
temp_dbfile = False


# Patch the tkintermapview CanvasPositionMarker 'calculate_text_y_offset()' function to
# allow labeling below the icon.  The icon anchor position is always "center" for this app.
def new_calc_text_offset(self):
	if self.icon is not None:
		if label_position == "below":
			self.text_y_offset = self.icon.height()/2 + 6 + label_size
		else:
			self.text_y_offset = -self.icon.height()/2 - 3
	else:
			self.text_y_offset = -56
tkmv.canvas_position_marker.CanvasPositionMarker.calculate_text_y_offset = new_calc_text_offset


#=====  Global constants and variables =====

# Tile servers for map basemap layers
bm_servers = {"OpenStreetMap": "https://a.tile.openstreetmap.org/{z}/{x}/{y}.png",
			"Google streets": "https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Google satellite": "https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Open topo map": "https://tile.opentopomap.org/{z}/{x}/{y}.png",
			"Stamen terrain": "https://stamen-tiles.a.ssl.fastly.net/terrain/{z}/{x}/{y}.png"
			#, "Stamen toner": "https://stamen-tiles.a.ssl.fastly.net/toner/{z}/{x}/{y}.png"
			}

# API keys for tile servers that require them.  The dictionary keys should match basemap names.
api_keys = {}

# Initial basemap to use
#initial_basemap = tuple(bm_servers.keys())[0]
initial_basemap = "OpenStreetMap"

# List of initial basemap names, for use when saving config
initial_bms = list(bm_servers.keys())


# X11 bitmaps for map icons
icon_xbm = {
	'ball': """#define ball_width 16
#define ball_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'bird': """#define bird_width 16
#define bird_height 16
static unsigned char bird.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x1c, 0x00, 0x3f, 0x80, 0xef, 0xc0, 0x7f, 0xe0, 0x3f,
   0xf0, 0x3f, 0xf8, 0x1f, 0xff, 0x1f, 0xfc, 0x0f, 0xe0, 0x07, 0x80, 0x01,
   0x00, 0x01, 0x00, 0x01, 0x80, 0x03, 0xe0, 0x0f};""",

	'block': """#define block_width 16
#define block_height 16
static unsigned char square_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0x00, 0x00, 0x00, 0x00};""",

	'bookmark': """#define bookmark_width 16
#define bookmark_height 16
static unsigned char bookmark_bits[] = {
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0x7e, 0x7e, 0x3e, 0x7c,
   0x1e, 0x78, 0x0e, 0x70, 0x06, 0x60, 0x02, 0x40};""",

	'circle': """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'diamond': """#define diamond_width 16
#define diamond_height 16
static unsigned char diamond_bits[] = {
   0x80, 0x01, 0xc0, 0x03, 0xe0, 0x07, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfc, 0x3f, 0xf8, 0x1f,
   0xf0, 0x0f, 0xe0, 0x07, 0xc0, 0x03, 0x80, 0x01};""",

   'fish': """#define fish_width 16
#define fish_height 16
static unsigned char fish.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x01, 0x00, 0x03, 0x01, 0x0f,
   0xe3, 0x7f, 0xf7, 0x78, 0x3e, 0xea, 0x9e, 0xfb, 0x73, 0x84, 0xc1, 0x63,
   0x01, 0x3e, 0x00, 0x18, 0x00, 0x18, 0x00, 0x08};""",

	'flag': """#define flag_width 16
#define flag_height 16
static unsigned char flag.xbm_bits[] = {
   0x00, 0x00, 0x0e, 0x00, 0x3e, 0x00, 0xfe, 0x01, 0xfe, 0x1f, 0xfe, 0xff,
   0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xf6, 0xff,
   0x86, 0xff, 0x06, 0xf8, 0x06, 0x00, 0x06, 0x00};""",

	'house': """#define house_width 16
#define house_height 16
static unsigned char house_bits[] = {
   0x80, 0x01, 0xc0, 0x33, 0x60, 0x36, 0xb0, 0x3d, 0xd8, 0x3b, 0xec, 0x37,
   0xf6, 0x6f, 0xfb, 0xdf, 0xfd, 0xbf, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e};""",

	'info': """#define info_width 16
#define info_height 16
static unsigned char info_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x7c, 0x3e, 0xfe, 0x7f, 0xfe, 0x7f,
   0x3f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7e, 0x7e, 0x7e, 0x7e,
   0x3c, 0x3c, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'lightning': """#define lightning_width 16
#define lightning_height 16
static unsigned char Lightning_bits[] = {
   0x00, 0xc0, 0x00, 0x70, 0x00, 0x1c, 0x00, 0x07, 0x80, 0x03, 0xe0, 0x01,
   0xf0, 0x00, 0xf8, 0x03, 0xc0, 0x3f, 0x00, 0x1f, 0x80, 0x07, 0xc0, 0x01,
   0xe0, 0x00, 0x38, 0x00, 0x0e, 0x00, 0x03, 0x00};""",

	'plus': """#define plus_width 16
#define plus_height 16
static unsigned char plus_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0xff, 0xff, 0xff, 0xff, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01};""",

	'rose': """#define rose_width 16
#define rose_height 16
static unsigned char rose_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xfc, 0x3f, 0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xe0, 0x07, 0xc0, 0x03,
   0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01, 0x80, 0x01};""",

	'square': """#define square_width 16
#define square_height 16
static unsigned char square_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",

	'star': """#define star_width 16
#define star_height 16
static unsigned char star_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f,
   0x7c, 0x3e, 0x3c, 0x3c, 0x0e, 0x70, 0x06, 0x60};""",

	'swamp': """#define swamp_width 16
#define swamp_height 16
static unsigned char swamp_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x01, 0x00, 0x01, 0x00, 0x09, 0x20, 0x09,
   0x20, 0x09, 0x20, 0x29, 0x24, 0x29, 0x24, 0x29, 0x24, 0x29, 0xe4, 0xaf,
   0xfd, 0xbc, 0x1d, 0xf0, 0x87, 0xc7, 0xf1, 0x9f};""",

	'target': """#define target_width 16
#define target_height 16
static unsigned char target_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x86, 0x61,
   0xc7, 0xe3, 0xe3, 0xc7, 0xe3, 0xc7, 0xc7, 0xe3, 0x86, 0x61, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'tree': """
#define tree_width 16
#define tree_height 16
static unsigned char tree_bits[] = {
   0xf8, 0x00, 0xa8, 0x37, 0x7c, 0x7d, 0xef, 0x4a, 0x37, 0xf5, 0xdf, 0xaf,
   0xbe, 0xdb, 0xfc, 0x7f, 0xb0, 0x77, 0xc0, 0x7b, 0xc0, 0x1f, 0xc0, 0x03,
   0xc0, 0x07, 0xc0, 0x07, 0xe0, 0x0f, 0xfc, 0x0f};""",

	'triangle': """#define triangle_width 16
#define triangle_height 16
static unsigned char triangle_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0xf0, 0x0f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'triangle_open': """#define triangle_open_width 16
#define triangle_open_height 16
static unsigned char triangle_open_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0x70, 0x0e, 0x70, 0x0e, 0x38, 0x1c, 0x38, 0x1c, 0x1c, 0x38, 0x1c, 0x38,
   0x0e, 0x70, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'wave': """#define wave_width 16
#define wave_height 16
static unsigned char wave_bits[] = {
   0x00, 0x00, 0x70, 0x00, 0xf8, 0x00, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x70, 0x3e, 0xf8, 0x1c, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x00};""",

	'wedge': """#define wedge_width 16
#define wedge_height 16
static unsigned char stn_marker_inv_bits[] = {
   0xff, 0xff, 0xff, 0x7f, 0xfe, 0x7f, 0xfe, 0x3f, 0xfc, 0x3f, 0xfc, 0x1f,
   0xf8, 0x1f, 0xf8, 0x0f, 0xf0, 0x0f, 0xf0, 0x07, 0xe0, 0x07, 0xe0, 0x03,
   0xc0, 0x03, 0xc0, 0x01, 0x80, 0x01, 0x80, 0x00};""",

	'whale': """#define whale.xbm_width 16
#define whale.xbm_height 16
static unsigned char whale.xbm_bits[] = {
   0x18, 0x00, 0x18, 0x00, 0x18, 0x00, 0x3f, 0x00, 0xf7, 0x00, 0xe0, 0x0f,
   0xc0, 0x0f, 0x80, 0x1f, 0x80, 0x3f, 0x00, 0x3f, 0x00, 0x7e, 0x30, 0x7c,
   0xf0, 0x6f, 0xc0, 0xdf, 0x00, 0xb0, 0x10, 0xe0};""",

	'x': """#define x_width 16
#define x_height 16
static unsigned char x_bits[] = {
   0x00, 0x00, 0x06, 0x60, 0x0e, 0x70, 0x1c, 0x38, 0x38, 0x1c, 0x70, 0x0e,
   0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0x70, 0x0e, 0x38, 0x1c,
   0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60, 0x00, 0x00};"""
	}

# X11 bitmaps for map button bar icons
expand_xbm = """#define expand_width 16
#define expand_height 16
static unsigned char expand_bits[] = {
   0x3f, 0xfc, 0x07, 0xe0, 0x0f, 0xf0, 0x1d, 0xb8, 0x39, 0x9c, 0x71, 0x8e,
   0x60, 0x06, 0x00, 0x00, 0x00, 0x00, 0x61, 0x06, 0x71, 0x8e, 0x39, 0x9c,
   0x1d, 0xb8, 0x0f, 0xf0, 0x07, 0xe0, 0x3f, 0xfc};"""

wedges_3_xbm = """#define wedges_3_width 16
#define wedges_3_height 16
static unsigned char wedges_3_bits[] = {
   0xff, 0x01, 0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00, 0x00, 0x00,
   0x80, 0xff, 0x00, 0x7f, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x08, 0xff, 0x01,
   0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00};"""

wedge_sm_xbm = """#define wedge_sm_width 16
#define wedge_sm_height 16
static unsigned char wedge_sm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xf8, 0x1f, 0xf8, 0x1f, 0xf0, 0x0f,
   0xf0, 0x0f, 0xe0, 0x07, 0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01,
   0x80, 0x01, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};"""

circle_xbm = """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};"""

cancel_xbm = """#define cancel_width 16
#define cancel_height 16
static unsigned char cancel_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x7c, 0x06, 0x6e,
   0x07, 0xe7, 0x83, 0xc3, 0xc3, 0xc1, 0xe7, 0xe0, 0x76, 0x60, 0x3e, 0x78,
   0x1c, 0x38, 0x78, 0x3e, 0xf0, 0x0f, 0xc0, 0x03};"""


# Color names for map symbols.  See https://www.w3schools.com/colors/colors_names.asp.
color_names = ("aliceblue", "antiquewhite", "aqua", "aquamarine", "azure", "beige", "bisque", "black", "blanchedalmond",
		"blue", "blueviolet", "brown", "burlywood", "cadetblue", "chartreuse", "chocolate", "coral", "cornflowerblue",
		"cornsilk", "crimson", "cyan", "darkblue", "darkcyan", "darkgoldenrod", "darkgray", "darkgrey", "darkgreen",
		"darkkhaki", "darkmagenta", "darkolivegreen", "darkorange", "darkorchid", "darkred", "darksalmon", "darkseagreen",
		"darkslateblue", "darkslategray", "darkslategrey", "darkturquose", "darkviolet", "deeppink", "deepskyblue",
		"dimgray", "dimgrey", "dodgerblue", "firebrick", "floralwhite", "forestgreen", "fuschia", "gainsboro", "ghostwhite",
		"gold", "goldenrod", "gray", "grey", "green", "greenyellow", "honeydew", "hotpink", "indianred", "indigo", "ivory",
		"khaki", "lavender", "lavenderblush", "lawngreen", "lemonchiffon", "lightblue", "lightcoral", "lightcyan",
		"lightgoldenrodyellow", "lightgray", "lightgrey", "lightgreen", "lightpink", "lightsalmon", "lightseagreen",
		"lightskyblue", "lightslategray", "lightslategrey", "lightsteelblue", "lightyellow", "lime", "limegreen", "linen",
		"magenta", "maroon", "mediumaquamarine", "mediumblue", "mediumorchid", "mediumpurple", "mediumseagreen",
		"mediumslateblue", "mediumspringgreen", "mediumturquose", "mediumvioletred", "midnightblue", "mintcream", "mistyrose",
		"moccasin", "navajowhite", "navy", "oldlace", "olive", "olivedrab", "orange", "orangered", "orchid", "palegoldenrod",
		"palegreen", "paleturquose", "palevioletred", "papayawhip", "peachpuff", "peru", "pink", "plum", "powderblue",
		"purple", "rebeccapurple", "red", "rosybrown", "royalblue", "saddlebrown", "salmon", "sandybrown", "seagreen",
		"seashell", "sienna", "silver", "skyblue", "slateblue", "slategray", "slategrey", "snow", "springgreen",
		"steelblue", "tan", "teal", "thistle", "tomato", "turquoise", "violet", "wheat", "white", "whitesmoke", "yellow",
		"yellowgreen")

# A shorter list for interactive selection of the marker color
select_colors = ('aqua', 'black', 'blue', 'blueviolet', 'brown', 'chartreuse', 'cornflowerblue', 'crimson',
		'cyan', 'darkblue', 'darkgreen', 'darkmagenta', 'darkorange', 'darkred', 'darkslategray', 'deeppink',
		'forestgreen', 'fuschia', 'green', 'greenyellow', 'magenta', 'maroon', 'navy', 'orange', 'orangered',
		'purple', 'red', 'violet', 'white', 'yellow', 'yellowgreen')


# List of imported symbol names and paths
imported_symbols = []

# Keys for custom symbols are made up of the color name and the icon name, separated with a space.
custom_icons = {}


# X11 bitmap for the application window icon
win_icon_xbm = """#define window_icon_width 16
#define window_icon_height 16
static unsigned char window_icon_bits[] = {
   0xff, 0xff, 0x01, 0x80, 0x01, 0x84, 0x01, 0x8e, 0x01, 0x9f, 0x81, 0xbf,
   0x21, 0x80, 0x71, 0x80, 0xf9, 0x80, 0xfd, 0x81, 0x01, 0x84, 0x01, 0x8e,
   0x01, 0x9f, 0x81, 0xbf, 0x01, 0x80, 0xff, 0xff};"""



#=====  Functions and classes  =====

def warning(message, kwargs):
	tk.messagebox.showerror("Warning", message, **kwargs)

def fatal_error(message, kwargs):
	tk.messagebox.showerror("Fatal error", message, **kwargs)
	sys.exit()


class CsvFile(object):
	def __init__(self, csvfname, junk_header_lines=0, dialect=None):
		self.csvfname = csvfname
		self.junk_header_lines = junk_header_lines
		self.lineformat_set = False		# Indicates whether delimiter, quotechar, and escapechar have been set
		self.delimiter = None
		self.quotechar = None
		self.escapechar = None
		self.inf = None
		self.colnames = None
		self.rows_read = 0
		# Python 3 only
		self.reader = csv.reader(open(csvfname, mode="rt", newline=''), dialect=dialect)
	def __next__(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		return row
	def next(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		if self.rows_read == 1:
			self.colnames = row
		return row
	def __iter__(self):
		return self


def treeview_sort_column(tv, col, reverse):
	# Sort columns in Tkinter Treeview.  From https://stackoverflow.com/questions/1966929/tk-treeview-column-sort#1967793
    colvals = [(tv.set(k, col), k) for k in tv.get_children()]
    colvals.sort(reverse=reverse)
    # Rearrange items in sorted positions
    for index, (val, k) in enumerate(colvals):
        tv.move(k, '', index)
    # Reverse sort next time
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))

def set_tv_headers(tvtable, column_headers, colwidths, charpixels):
	pixwidths = [charpixels * col for col in colwidths]
	for i in range(len(column_headers)):
		hdr = column_headers[i]
		tvtable.column(hdr, width=pixwidths[i])
		tvtable.heading(hdr, text=hdr, command=lambda _col=hdr: treeview_sort_column(tvtable, _col, False))

def fill_tv_table(tvtable, rowset, status_label=None):
	for i, row in enumerate(rowset):
		enc_row = [c if c is not None else '' for c in row]
		tvtable.insert(parent='', index='end', iid=str(i), values=enc_row)
	if status_label is not None:
		status_label.config(text = "    %d rows" % len(rowset))

def treeview_table(parent, rowset, column_headers, select_mode="none"):
	# Creates a TreeView table containing the specified data, with scrollbars and status bar
	# in an enclosing frame.
	# This does not grid the table frame in its parent widget.
	# Returns a tuple of 0: the frame containing the table,  and 1: the table widget itself.
	nrows = range(len(rowset))
	ncols = range(len(column_headers))
	hdrwidths = [len(column_headers[j]) for j in ncols]
	if len(rowset) > 0:
		if sys.version_info < (3,):
			datawidthtbl = [[len(rowset[i][j] if isinstance(rowset[i][j], str) else type(u"")(rowset[i][j])) for i in nrows] for j in ncols]
		else:
			datawidthtbl = [[len(rowset[i][j] if isinstance(rowset[i][j], str) else str(rowset[i][j])) for i in nrows] for j in ncols]
		datawidths = [max(cwidths) for cwidths in datawidthtbl]
	else:
		datawidths = hdrwidths
	colwidths = [max(hdrwidths[i], datawidths[i]) for i in ncols]
	# Set the font.
	ff = tkfont.nametofont("TkFixedFont")
	tblstyle = ttk.Style()
	tblstyle.configure('tblstyle', font=ff)
	charpixels = int(1.3 * ff.measure(u"0"))
	tableframe = ttk.Frame(master=parent, padding="3 3 3 3")
	statusframe = ttk.Frame(master=tableframe)
	# Create and configure the Treeview table widget
	tv_widget = ttk.Treeview(tableframe, columns=column_headers, selectmode=select_mode, show="headings")
	tv_widget.configure()["style"] = tblstyle
	ysb = ttk.Scrollbar(tableframe, orient='vertical', command=tv_widget.yview)
	xsb = ttk.Scrollbar(tableframe, orient='horizontal', command=tv_widget.xview)
	tv_widget.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
	# Status bar
	parent.statusbar = ttk.Label(statusframe, text="    %d rows" % len(rowset), relief=tk.RIDGE, anchor=tk.W)
	tableframe.statuslabel = parent.statusbar
	# Fill the Treeview table widget with data
	set_tv_headers(tv_widget, column_headers, colwidths, charpixels)
	fill_tv_table(tv_widget, rowset, parent.statusbar)
	# Place the table
	tv_widget.grid(column=0, row=0, sticky=tk.NSEW)
	ysb.grid(column=1, row=0, sticky=tk.NS)
	xsb.grid(column=0, row=1, sticky=tk.EW)
	statusframe.grid(column=0, row=3, sticky=tk.EW)
	tableframe.columnconfigure(0, weight=1)
	tableframe.rowconfigure(0, weight=1)
	# Place the status bar
	parent.statusbar.pack(side=tk.BOTTOM, fill=tk.X)
	#
	return tableframe, tv_widget

def dquote(v):
	# Returns a double-quoted value if it is not an identifier.
	if not v.isidentifier():
		return '"%s"' % v
	else:
		return v

def db_colnames(tbl_hdrs):
	# Takes a list of table headers and returns a list of database column names,
	# with doubl-quoting of any name that is not all alphanumeric and starts with
	# an alphabetic.
	colnames = []
	for hdr in tbl_hdrs:
		colnames.append(dquote(hdr))
	return colnames

def isint(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	if type(v) == int:
		return True
	if type(v) == float:
		return False
	try:
		int(v)
		return True
	except ValueError:
		return False

def isfloat(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		float(v)
		return True
	except ValueError:
		return False

def dt_type(v):
	# Type of date/time: timestamp, date, or None
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	t = None
	try:
		t = dateparser.isoparse(v)
	except:
		try:
			t = dateparser.parse(v, ignoretz=True)
		except:
			pass
	if t is None:
		return None
	if t.time() == datetime.time(0,0):
		return "date"
	else:
		return "timestamp"

def data_type(v):
	# Characterizes the value v as one of a simple set of data types.
	# Returns "timestamp", "date", "int", "float", or "string"
	if v is None or (type(v) is str and v == ''):
		return None
	if isint(v):
		return "int"
	if isfloat(v):
		return "float"
	dt = dt_type(v)
	if dt is not None:
		return dt
	return "string"

def data_type_cast_fn(data_type_str):
	if data_type_str == "string":
		return str
	elif data_type_str == "date":
		return datetime.date
	elif data_type_str == "timestamp":
		return datetime.timestamp
	elif data_type_str == "int":
		return int
	elif data_type_str == "float":
		return float

def common_data_type(values):
	# Returns a data type common to all the values in the list.
	# This is "string" unlees all types are the same.
	# Null (None) values are ignored.  If all values are null, return "string".
	val2 = [v for v in values if v is not None and not (type(v) is str and v.strip() == '')]
	if len(val2) == 0:
		return "string"
	else:
		types = [data_type(v) for v in val2]
		if len(set(types)) == 1:
			return types[0]
		else:
			return "string"

def set_data_types(headers, rows, data_type_list):
	return_queue = queue.Queue()
	def eval_columns(headers, rows, return_queue):
		coltypes = []
		for i, colname in enumerate(headers):
			datavals = [row[i] for row in rows]
			non_null = [d for d in datavals if d is not None and not (type(d) is str and d.strip() == '')]
			nullcount = len(datavals) - len(non_null)
			uniquevals = len(set(non_null))
			coltypes.append((colname, common_data_type(datavals), nullcount, uniquevals))
		return_queue.put(coltypes)
	t = threading.Thread(target=eval_columns, args=(headers, rows, return_queue))
	t.start()
	return (t, return_queue)

# Translations to SQLite type affinity names
sqlite_type_x = {'int': 'INTEGER', 'float': 'REAL', 'string': 'TEXT', 'timestamp': 'TEXT', 'date': 'TEXT'}

def center_window(win):
	win.update_idletasks()
	m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", win.geometry())
	if m is not None:
		wwd = int(m.group(1))
		wht = int(m.group(2))
		swd = win.winfo_screenwidth()
		sht = win.winfo_screenheight()
		xpos = (swd/2) - (wwd/2)
		ypos = (sht/2) - (wht/2)
		win.geometry("%dx%d+%d+%d" % (wwd, wht, xpos, ypos))

def raise_window(win):
	win.attributes('-topmost', 1)
	win.attributes('-topmost', 0)



class MapUI(object):
	def __init__(self, src_name, message, lat_col, lon_col, crs=4326, sheet=None,
			label_col=None, symbol_col=None, color_col=None,
			map_export_file=None, export_time_sec=10):
		self.win = tk.Tk()
		if src_name is None:
			self.win._root().withdraw()
			sdsd = SelDataSrcDialog()
			src_name, label_col, lat_col, lon_col, crs, symbol_col, color_col, message, headers, rows = sdsd.select()
			if src_name is None:
				self.cancel()
			self.win._root().deiconify()
		else:
			# src_name is a filename, either CSV or spreadsheet
			fn, ext = os.path.splitext(src_name)
			if ext.lower() == ".csv":
				try:
					headers, rows = file_data(src_name)
				except:
					self.win._root().withdraw()
					fatal_error("Could not read data from %s" % src_data, kwargs={'parent': self.win})
			else:
				if sheet is None:
					self.win._root().withdraw()
					fatal_error("A sheet name must be specified for spreadsheets", kwargs={'parent': self.win})
				try:
					if ext.lower() == '.ods':
						headers, rows = ods_data(src_name, sheet)
					else:
						headers, rows = xls_data(src_name, sheet)
				except:
					self.win._root().withdraw()
					fatal_error("Could not read table from %s, sheet %s" % (src_name, sheet), kwargs={'parent': self.win})
		self.win.protocol("WM_DELETE_WINDOW", self.cancel)
		self.data_src_name = src_name
		self.win.title("Map of %s" % src_name)
		self.headers = headers
		self.rows = rows
		# Set the font
		self.mapfont = self.makefont()
		# Set the application window icon
		#win_icon = tk.BitmapImage(data=win_icon_xbm, foreground="black", background="tan")
		#self.win.iconbitmap(win_icon)
		# Source and possibly un-projected crs
		self.src_crs = crs
		self.crs = crs
		# Created column names for un-projected coordinates
		self.lat_4326_col = None
		self.lon_4326_col = None
		# The markers for all the locations in the data table
		self.loc_map_markers = []
		# The markers for the selected location(s)
		self.sel_map_markers = []
		# The number of table rows without coordinates
		self.missing_latlon = 0
		# Map bounds
		self.min_lat = None
		self.max_lat = None
		self.min_lon = None
		self.max_lon = None
		# List of PlotDialog objects, so they can have data pushed or be deleted.
		self.plot_list = []
		# Database connection is set in 'add_data()'; variables are initialized here
		self.dbtmpdir = None
		self.dbname = None
		self.db = None
		# Create default markers for the map
		self.loc_marker_icon = self.set_get_loc_marker()
		# Initializes selection marker to the global settings
		self.set_sel_marker(select_symbol, select_color)
		# Create icons for the buttonbar
		expand_icon = tk.BitmapImage(data=expand_xbm, foreground="black")
		focus_icon = tk.BitmapImage(data=wedge_sm_xbm, foreground="red")
		zoom_sel_icon = tk.BitmapImage(data=wedges_3_xbm, foreground="red")
		unselect_icon = tk.BitmapImage(data=cancel_xbm, foreground="black")
		# Use stacked frames for the main application window components.  Map and table in a PanedWindow.
		msgframe = ttk.Frame(self.win, padding="3 2")
		ctrlframe = ttk.Frame(self.win, padding="3 2")
		datapanes = ttk.PanedWindow(self.win, orient=tk.VERTICAL)
		mapframe = ttk.Frame(datapanes, borderwidth=2, relief=tk.RIDGE)
		self.tblframe = ttk.Frame(datapanes, padding="3 2")
		datapanes.add(mapframe, weight=1)
		datapanes.add(self.tblframe, weight=1)
		# Allow vertical resizing of map and table frames, not of message and control frames
		self.win.columnconfigure(0, weight=1)
		self.win.rowconfigure(0, weight=0)		# msgframe
		self.win.rowconfigure(1, weight=0)		# ctrlframe
		self.win.rowconfigure(2, weight=1)		# datapanes
		# Grid all the main frames
		msgframe.grid(row=0, column=0, sticky=tk.NSEW)
		ctrlframe.grid(row=1, column=0, sticky=tk.W)
		datapanes.grid(row=2, column=0, sticky=tk.NSEW)
		# Populate the message frame
		self.msg_label = ttk.Label(msgframe, text=message)
		def wrap_msg(event):
			self.msg_label.configure(wraplength=event.width - 5)
		self.msg_label.bind("<Configure>", wrap_msg)
		self.msg_label.grid(column=0, row=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		msgframe.columnconfigure(0, weight=1)
		msgframe.rowconfigure(0, weight=1)
		# Populate the map control frame
		ctrlframe.rowconfigure(0, weight=0)
		ctrlframe.columnconfigure(0, weight=0)
		# Basemap controls and buttons
		self.basemap_label = ttk.Label(ctrlframe, text="Basemap:", anchor="w")
		self.basemap_label.grid(row=0, column=0, padx=(5, 5), pady=(2, 5), sticky=tk.W)
		global initial_basemap
		bm_name = initial_basemap
		if bm_name not in bm_servers:
			bm_name = tuple(bm_servers.keys())[0]
			initial_basemap = bm_name
		self.basemap_var = tk.StringVar(self.win, bm_name)
		self.map_option_menu = ttk.Combobox(ctrlframe, state="readonly", textvariable=self.basemap_var,
				values=self.available_tile_servers(), width=18)
		self.map_option_menu.bind('<<ComboboxSelected>>', self.change_basemap)
		self.map_option_menu.grid(row=0, column=1, padx=(5, 30), pady=(2, 5), sticky=tk.W)
		# Multi-select option
		def ck_changed():
			ck = self.multiselect_var.get()
			if ck == '0':
				self.unselect_map()
				self.tbl.configure(selectmode = tk.BROWSE)
			else:
				self.tbl.configure(selectmode = tk.EXTENDED)
			self.set_status()
		# Set by global variable
		self.multiselect_var = tk.StringVar(self.win, multiselect)
		ck_multiselect = ttk.Checkbutton(ctrlframe, text="Multi-select", variable=self.multiselect_var, command=ck_changed)
		ck_multiselect.grid(row=0, column=2, sticky=tk.W, padx=(0, 20))
		# Map control buttons
		zoomsel_btn = ttk.Button(ctrlframe, text="Zoom selected", image=zoom_sel_icon, compound=tk.LEFT, command=self.zoom_selected)
		zoomsel_btn.image = zoom_sel_icon
		zoomsel_btn.grid(row=0, column=3, sticky=tk.W)
		expand_btn = ttk.Button(ctrlframe, text="Zoom full", image=expand_icon, compound=tk.LEFT, command=self.zoom_full)
		expand_btn.image = expand_icon
		expand_btn.grid(row=0, column=4, sticky=tk.W)
		focus_btn = ttk.Button(ctrlframe, text="Center", image=focus_icon, compound=tk.LEFT, command=self.focus_map)
		focus_btn.image = focus_icon
		focus_btn.grid(row=0, column=5, sticky=tk.W)
		unselect_btn = ttk.Button(ctrlframe, text="Un-select", image=unselect_icon, compound=tk.LEFT, command=self.unselect_map)
		unselect_btn.image = unselect_icon
		unselect_btn.grid(row=0, column=6, sticky=tk.W)
		# Map widget
		mapframe.rowconfigure(0, weight=1)
		mapframe.columnconfigure(0, weight=1)
		self.map_widget = tkmv.TkinterMapView(mapframe, height=600, width=600, corner_radius=0)
		if initial_basemap != "OpenMapServer":
			tileserver = self.tile_url(initial_basemap)
			self.map_widget.set_tile_server(tileserver)
		self.map_widget.grid(row=0, column=0, sticky=tk.NSEW)
		# Populate the table frame
		self.tblframe.rowconfigure(0, weight=1)
		self.tblframe.columnconfigure(0, weight=1)
		self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col,
				symbol_col, color_col)
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
		# Add menu
		self.add_menu(table_object = self.tbl, column_headers=headers)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
		# Other key bindings
		self.win.protocol("WM_DELETE_WINDOW", self.cancel)
		# Position window.
		self.win.update_idletasks()
		m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", self.win.geometry())
		if m is not None:
			wwd = int(m.group(1))
			wht = int(m.group(2))
			swd = self.win.winfo_screenwidth()
			sht = self.win.winfo_screenheight()
			xpos = (swd/2) - (wwd/2)
			ypos = (sht/2) - (wht/2)
			self.win.geometry("%dx%d+%d+%d" % (wwd, wht, xpos, ypos))
		# Limit resizing
		self.win.minsize(width=400, height=400)
		# Set table status message
		self.set_status()
		# Just export the map and quit?
		if map_export_file is not None:
			self.imageoutputfile = map_export_file
			self.win.after(export_time_sec * 1000, self.export_map_and_quit)
	def makefont(self):
		global label_font, label_size, label_bold
		fams = tkfont.families()
		if not label_font in fams:
			alt_fonts = ["Liberation Sans", "Arial", "Helvetica", "Nimbus Sans", "Liberation Sans", "Trebuchet MS", "Tahoma", "DejaVu Sans", "Bitstream Vera Sans", "Open Sans"]
			font_found = False
			for f in alt_fonts:
				if f in fams:
					label_font = f
					font_found = True
					break
			if not font_found:
				label_font = tkfont.nametofont("TkDefaultFont").actual()["family"]
		boldstr = "normal" if not label_bold else "bold"
		return tkfont.Font(family=label_font, size=label_size, weight=boldstr)
	def available_tile_servers(self):
		# Return a list of those without API keys or for which API keys are provided
		avail = []
		for k in bm_servers:
			if self.tile_url(k) is not None:
				avail.append(k)
		return avail
	def tile_url(self, source_name):
		# Return the URL with the API key replaced, unless it is not available.
		source_url = bm_servers[source_name]
		if "<api_key>" in source_url.lower():
			if source_name in api_keys:
				api_key = api_keys[source_name]
				for matched in re.findall("<api_key>", source_url, re.IGNORECASE):
					source_url = source_url.replace(matched, api_key)
				return source_url
			else:
				return None
		else:
			return source_url
	def mark_map(self, event):
		# Highlight the selected row(s) in the table and get the coordinates to map it
		if self.tbl.selection():
			new_markers = []
			for sel_row in self.tbl.selection():
				rowdata = self.tbl.item(sel_row)["values"]
				try:
					lat_val = float(rowdata[self.lat_index])
				except:
					lat_val = None
				try:
					lon_val = float(rowdata[self.lon_index])
				except:
					lon_val = None
				if lon_val is not None and lat_val is not None:
					new_marker = self.map_widget.set_marker(lat_val, lon_val, icon=self.sel_marker_icon)
					new_markers.append(new_marker)
			for m in self.sel_map_markers:
				m.delete()
			self.sel_map_markers = new_markers
		else:
			for m in self.sel_map_markers:
				m.delete()
		self.update_plot_data()
		self.set_status()
	def set_sel_marker(self, symbol, color):
		select_marker = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
		mkr_key = "%s %s" % (color, symbol)
		if mkr_key not in custom_icons:
			custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
		select_marker = custom_icons[mkr_key]
		self.sel_marker_icon = select_marker
	def redraw_sel_markers(self):
		new_markers = []
		for mkr in self.sel_map_markers:
			mposition = mkr.position
			micon = mkr.icon
			mkr.delete()
			new_marker = self.map_widget.set_marker(mposition[0], mposition[1], icon=micon)
			new_markers.append(new_marker)
		self.sel_map_markers = new_markers
	def draw_sel_markers(self):
		for mkr in self.sel_map_markers:
			mkr.draw()
	def set_get_loc_marker(self):
		mkr_key = "%s %s" % (location_color, location_marker)
		if mkr_key not in custom_icons:
			custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[location_marker], foreground=location_color)
		return custom_icons[mkr_key]
	def redraw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.draw_loc_markers(tdata)
	def draw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		# Also set the number of rows missing coordinates and the bounding box.
		self.missing_latlon = 0
		for row_id in tdata.get_children():
			rowdata = tdata.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if self.min_lat is None or lat_val < self.min_lat:
					self.min_lat = lat_val
				if self.max_lat is None or lat_val > self.max_lat:
					self.max_lat = lat_val
				if self.min_lon is None or lon_val < self.min_lon:
					self.min_lon = lon_val
				if self.max_lon is None or lon_val > self.max_lon:
					self.max_lon = lon_val
				if self.color_index is None and self.symbol_index is None:
					marker_icon = self.loc_marker_icon
				else:
					if self.color_index is None or not use_data_color:
						color = location_color
					else:
						color = rowdata[self.color_index].lower()
						if color not in color_names:
							color = location_color
					if self.symbol_index is None or not use_data_marker:
						symbol = location_marker
					else:
						symbol = rowdata[self.symbol_index].lower()
						if symbol not in icon_xbm:
							symbol = location_marker
					mkr_key = "%s %s" % (color, symbol)
					if mkr_key not in custom_icons:
						custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
					marker_icon = custom_icons[mkr_key]
				if self.label_index is not None:
					lbl = rowdata[self.label_index]
					mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon,
							text=lbl, font=self.mapfont, text_color=label_color,
							command=self.map_sel_table)
					self.loc_map_markers.append(mkr)
				else:
					mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon, command=self.map_sel_table)
					self.loc_map_markers.append(mkr)
			else:
				self.missing_latlon += 1
		self.update_plot_data()
	def add_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		# Launch separate process to determine data types
		self.data_types = []
		(dt_thread, dt_queue) = set_data_types(headers, rows, self.data_types)
		# Re-set all data-specific variables and widgets
		self.lat_col = lat_col
		self.lon_col = lon_col
		self.src_lat_col = lat_col
		self.src_lon_col = lon_col
		self.label_col = label_col
		self.symbol_col = symbol_col
		self.color_col = color_col
		self.lat_index = headers.index(lat_col)
		self.lon_index = headers.index(lon_col)
		self.src_lat_index = headers.index(lat_col)
		self.src_lon_index = headers.index(lon_col)
		self.label_index = headers.index(label_col) if label_col is not None and label_col != '' else None
		self.symbol_index = headers.index(symbol_col) if symbol_col is not None and symbol_col != '' else None
		self.color_index = headers.index(color_col) if color_col is not None and color_col != '' else None
		if self.crs != 4326:
			try:
				from pyproj import CRS, Transformer
			except:
				fatal_error("The pyproj library is required to re-project spatial coordinates")
			try:
				crs_proj = CRS(self.crs)
			except:
				fatal_error("Invalid CRS (%s)" % self.crs)
			if self.lat_4326_col is None:
				for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
					if colname not in headers:
						self.lat_4326_col = colname
						headers.append(colname)
						break
			if self.lon_4326_col is None:
				for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
					if colname not in headers:
						self.lon_4326_col = colname
						headers.append(colname)
						break
			self.lat_col = self.lat_4326_col
			self.lon_col = self.lon_4326_col
			crs_4326 = CRS(4326)
			reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
			for r in rows:
				y = r[self.src_lat_index]
				x = r[self.src_lon_index]
				if y is not None and y != 0 and x is not None and x != 0:
					try:
						newx, newy = reproj.transform(x, y)
						r.extend([newy, newx])
					except:
						r.extend([None, None])
				else:
					r.extend([None, None])
			self.lat_index = headers.index(self.lat_col)
			self.lon_index = headers.index(self.lon_col)
		tframe, tdata = treeview_table(self.tblframe, rows, headers, "browse")
		self.table_row_count = len(tdata.get_children())
		# Scan the table, put points on the map, and find the map extent.
		self.min_lat = self.max_lat = self.min_lon = self.max_lon = None
		self.sel_map_markers = []
		self.missing_latlon = 0
		self.draw_loc_markers(tdata)
		# Set the map extent based on coordinates.
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))
		# Copy data from the treeview table to the database.  This includes the treeview IDs
		# Database connection
		if self.db is not None:
			self.db.close()
		if self.dbtmpdir is not None:
			self.dbtmpdir.cleanup(ignore_cleanup_errors = True)
		if temp_dbfile:
			self.dbtmpdir = tempfile.TemporaryDirectory()
			self.dbname = os.path.join(self.dbtmpdir.name, "mapdata.db")
			self.db = sqlite3.connect(self.dbname)
		else:
			self.tmpdir = None
			self.dbname = None
			self.db = sqlite3.connect(":memory:")
		cur = self.db.cursor()
		colnames = db_colnames(headers)
		colnames.append("treeviewid")
		cur.execute("create table mapdata (%s);" % ",".join(colnames))
		tbldata = []
		for row_id in tdata.get_children():
			row_vals = tdata.item(row_id)["values"]
			row_vals.append(row_id)
			tbldata.append(row_vals)
		params = ",".join(['?'] * len(colnames))
		cur.executemany("insert into mapdata values (%s)" % params, tbldata)
		cur.close()
		# Initial value for user-entered WHERE clause
		self.whereclause = ""
		# Save data types for use in column selection for plotting
		dt_thread.join()
		self.data_types = dt_queue.get(block=True)
		# Return frame and data table
		return tframe, tdata
	def remove_data(self):
		while len(self.sel_map_markers) > 0:
			self.sel_map_markers.pop().delete()
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.map_widget.delete_all_marker()
		self.close_all_plots()
		self.tableframe.destroy()
		self.tbl.destroy()
	def set_tbl_selectmode(self):
		ck = self.multiselect_var.get()
		if ck == '0':
			self.tbl.configure(selectmode = tk.BROWSE)
		else:
			self.tbl.configure(selectmode = tk.EXTENDED)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
	def replace_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		self.remove_data()
		self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col, symbol_col, color_col)
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
	def new_data_file(self):
		dfd = DataFileDialog()
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_spreadsheet_file(self):
		dfd = ImportSpreadsheetDialog()
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_db_table(self):
		dbd = DbConnectDialog()
		tablename, id_col, lat_col, lon_col, crs, sym_col, col_col, desc, headers, rows = dbd.get_data()
		if tablename is not None and tablename != '':
			self.crs = crs
			self.win.title("Map of %s" % tablename)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if desc is not None and desc != '':
				self.msg_label['text'] = desc
	def zoom_full(self):
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))
	def zoom_selected(self):
		if len(self.sel_map_markers) > 0:
			if len(self.sel_map_markers) == 1:
				self.focus_map()
			else:
				min_lat = max_lat = min_lon = max_lon = None
				for m in self.sel_map_markers:
					lat, lon = m.position
					if min_lat is None or lat < min_lat:
						min_lat = lat
					if max_lat is None or lat > max_lat:
						max_lat = lat
					if min_lon is None or lon < min_lon:
						min_lon = lon
					if max_lon is None or lon > max_lon:
						max_lon = lon
			self.map_widget.fit_bounding_box((max_lat, min_lon), (min_lat, max_lon))
	def focus_map(self):
		# Center the map on the last marker
		if len(self.sel_map_markers) > 0:
			m = self.sel_map_markers[-1]
			self.map_widget.set_position(m.position[0], m.position[1])
	def unselect_map(self):
		for m in self.sel_map_markers:
			self.map_widget.delete(m)
		self.tbl.selection_remove(*self.tbl.selection())
		self.sel_map_markers = []
		self.update_plot_data()
		self.set_status()
	def change_basemap(self, *args):
		new_map = self.basemap_var.get()
		tileserver = self.tile_url(new_map)
		self.map_widget.set_tile_server(tileserver)
	def map_sel_table(self, marker):
		# Highlight the table row for the clicked map marker
		lat, lon = marker.position
		if self.multiselect_var.get() == '0':
			for mkr in self.sel_map_markers:
				self.map_widget.delete(mkr)
			self.sel_map_markers = []
			self.tbl.selection_remove(*self.tbl.selection())
		for row_id in self.tbl.get_children():
			rowdata = self.tbl.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if lat_val == lat and lon_val == lon:
					self.tbl.selection_add(row_id)
					self.tbl.see(row_id)
					new_marker = self.map_widget.set_marker(lat, lon, icon=self.sel_marker_icon)
					if not new_marker in self.sel_map_markers:
						self.sel_map_markers.append(new_marker)
		self.update_plot_data()
		self.set_status()
	def set_status(self):
		statusmsg = "    %d rows" % self.table_row_count
		if self.missing_latlon > 0:
			statusmsg = statusmsg + " (%d without lat/lon)" % self.missing_latlon
		if len(self.tbl.selection()) > 0:
			statusmsg = statusmsg + "  |  %s selected" % len(self.tbl.selection())
		if self.multiselect_var.get() == "1":
			statusmsg = statusmsg + "  |  Ctrl-click to select multiple rows"
		self.tblframe.statusbar.config(text = statusmsg)

	def get_all_data(self, column_list):
		# Plotting support.  Return all data for the specified columns as a list of column-oriented lists.
		res = []
		for c in column_list:
			i = self.headers.index(c)
			res.append([row[i] for row in self.rows])
		return res
	def get_sel_data(self, column_list):
		# Plotting support.  Return data from selected rows for the specified columns, as a list of lists.
		res = [[] for _ in column_list]
		indices = [self.headers.index(c) for c in column_list]
		for sel_row in self.tbl.selection():
			datarow = self.tbl.item(sel_row)["values"]
			for i, index in enumerate(indices):
				res[i].append(datarow[index])
		return res
	def update_plot_data(self):
		for plot in self.plot_list:
			if plot.sel_only_var.get() == "1" and plot.auto_update:
				plot.q_redraw()
	def remove_plot(self, plot_obj):
		# For use by the plot 'do_close()' method.
		try:
			self.plot_list.remove(plot_obj)
		except:
			pass
	def close_plot(self, plot_obj):
		try:
			plot_obj.do_close()
			self.remove_plot()
		except:
			pass
	def close_all_plots(self):
		while len(self.plot_list) > 0:
			self.plot_list[0].do_close()
			# The callback will remove the plot.
		self.plot_list = []

	def change_crs(self):
		crsdlg = NewCrsDialog(self.crs)
		new_crs = crsdlg.get_crs()
		if new_crs is not None:
			if new_crs != self.crs:
				try:
					from pyproj import CRS, Transformer
				except:
					fatal_error("The pyproj library is required to re-project spatial coordinates")
				try:
					crs_proj = CRS(new_crs)
				except:
					warning("Invalid CRS (%s)" % new_crs, kwargs={})
				else:
					if self.lat_4326_col is None:
						for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
							if colname not in self.headers:
								self.lat_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					if self.lon_4326_col is None:
						for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
							if colname not in self.headers:
								self.lon_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					self.lat_col = self.lat_4326_col
					self.lon_col = self.lon_4326_col
					self.lat_index = self.headers.index(self.lat_4326_col)
					self.lon_index = self.headers.index(self.lon_4326_col)
					crs_4326 = CRS(4326)
					self.crs = new_crs
					reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
					for r in self.rows:
						y = r[self.src_lat_index]
						x = r[self.src_lon_index]
						if y is not None and y != 0 and x is not None and x != 0:
							try:
								newx, newy = reproj.transform(x, y)
								r[self.lat_index] = newy
								r[self.lon_index] = newx
							except:
								r[self.lat_index] = None
								r[self.lon_index] = None
						else:
							r[self.lat_index] = None
							r[self.lon_index] = None
					selected = self.tbl.selection()
					self.replace_data(self.rows, self.headers, self.src_lat_col, self.src_lon_col, self.label_col, self.symbol_col, self.color_col)
					self.tbl.selection_set(tuple(selected))
					self.mark_map({})
	def cancel(self):
		self.win.destroy()
		sys.exit()
	def export_map_and_quit(self):
		fn, ext = os.path.splitext(self.imageoutputfile)
		if ext.lower() == ".ps":
			self.export_map_to_ps(self.imageoutputfile)
		else:
			self.map_widget.update_idletasks()
			#self.win.after(200, self.save_imageoutputfile)
			self.save_imageoutputfile()
		self.win.destroy()
	def export_map_to_ps(self, outfile):
		self.map_widget.canvas.postscript(file=outfile, colormode='color')
	def save_imageoutputfile(self):
		obj = self.map_widget.canvas
		bounds = (obj.winfo_rootx(), obj.winfo_rooty(), 
				obj.winfo_rootx() + obj.winfo_width(), obj.winfo_rooty() + obj.winfo_height())
		ImageGrab.grab(bbox=bounds).save(self.imageoutputfile)
	def export_map_to_img(self, outfile):
		# Allow map to recover from blocking by the file dialog box before grabbing and exporting the canvas
		self.map_widget.update_idletasks()
		self.imageoutputfile = outfile
		self.win.after(1000, self.save_imageoutputfile)
	def add_menu(self, table_object, column_headers):
		mnu = tk.Menu(self.win)
		self.win.config(menu=mnu)
		file_menu = tk.Menu(mnu, tearoff=0)
		tbl_menu = tk.Menu(mnu, tearoff=0)
		map_menu = tk.Menu(mnu, tearoff=0)
		sel_menu = tk.Menu(mnu, tearoff=0)
		plot_menu = tk.Menu(mnu, tearoff=0)
		help_menu = tk.Menu(mnu, tearoff=0)
		mnu.add_cascade(label="File", menu=file_menu, underline=0)
		mnu.add_cascade(label="Table", menu=tbl_menu, underline=0)
		mnu.add_cascade(label="Map", menu=map_menu, underline=0)
		mnu.add_cascade(label="Selections", menu=sel_menu, underline=0)
		mnu.add_cascade(label="Plot", menu=plot_menu, underline=0)
		mnu.add_cascade(label="Help", menu=help_menu, underline=0)
		def save_table():
			if table_object.selection():
				rowset = []
				for sel_row in table_object.selection():
					rowset.append(table_object.item(sel_row)["values"])
				outfile = tkfiledialog.asksaveasfilename(title="File to save selected rows",
					filetypes=[('CSV files', '.csv'), ('ODS files', '.ods'), ('TSV files', '.tsv'), ('Plain text', '.txt'), ('LaTeX', '.tex')])
				if outfile:
					if outfile[-3:].lower() == 'csv':
						write_delimited_file(outfile, "csv", column_headers, rowset)
					elif outfile[-3:].lower() == 'tsv':
						write_delimited_file(outfile, "tsv", column_headers, rowset)
					elif outfile[-3:].lower() == 'txt':
						write_delimited_file(outfile, "plain", column_headers, rowset)
					elif outfile[-3:].lower() == 'tex':
						write_delimited_file(outfile, "tex", column_headers, rowset)
					elif outfile[-3:].lower() == 'ods':
						export_ods(outfile, column_headers, rowset, append=True, sheetname="Selected map items")
					else:
						# Force write as CSV.
						outfile = outfile + ".csv"
						write_delimited_file(outfile, "csv", column_headers, rowset)
		def save_map():
			outfile = tkfiledialog.asksaveasfilename(title="File to save map",
				filetypes=[('Postscript files', '.ps'), ('JPEG files', '.jpg'), ('PNG files', '.png')])
			fn, ext = os.path.splitext(outfile)
			if len(ext) > 1 and outfile[-2:].lower() == 'ps':
				self.export_map_to_ps(outfile)
			else:
				self.export_map_to_img(outfile)
		def change_marker():
			global select_symbol, select_color
			marker_dlg = MarkerDialog(map_menu)
			symbol, color = marker_dlg.get_marker()
			if symbol is not None or color is not None:
				if symbol is None or symbol == '':
					symbol = select_symbol
				if color is None or color == '':
					color = select_color
				symb_name = "%s %s" % (color, symbol)
				if symb_name not in custom_icons:
					custom_icons[symb_name] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
				select_symbol = symbol
				select_color = color
				self.sel_marker_icon = custom_icons[symb_name]
		def change_labeling():
			global location_marker, location_color, use_data_marker, use_data_color
			global label_font, label_size, label_bold, label_position
			lbl_dlg = LabelDialog(map_menu, self.headers, self.label_col)
			mkr, clr, datamkr, dataclr, column, ffam, fsize, fbold, pos = lbl_dlg.get_labeling()
			fsize = int(fsize)
			fbold = False if "0" else True
			if mkr is not None:
				if mkr != location_marker or clr != location_color or datamkr != use_data_marker or \
						dataclr != use_data_color or column != self.label_col or ffam != label_font or \
						fsize != label_size or fbold != label_bold or pos != label_position:
							fontchanged = ffam != label_font or fsize != label_size or fbold != label_bold
							location_marker = mkr
							location_color = clr
							use_data_marker = datamkr
							use_data_color = dataclr
							self.label_col = column if column != '' else None
							label_font = ffam
							label_size = fsize
							label_bold = fbold
							label_position = pos
							if fontchanged:
								self.mapfont = self.makefont()
							self.loc_marker_icon = self.set_get_loc_marker()
							self.label_index = self.headers.index(self.label_col) if self.label_col is not None and self.label_col != '' else None
							self.redraw_loc_markers(self.tbl)
							self.redraw_sel_markers()
		def import_symbol_file():
			sd = ImportSymbolDialog()
			name, fn = sd.run()
			if name is not None and fn is not None:
				import_symbol(name, fn)
				fqfn = os.path.abspath(fn)
				symb_spec = (name, fqfn)
				if not symb_spec in imported_symbols:
					imported_symbols.append(symb_spec)
		def read_config_file():
			fn = tkfiledialog.askopenfilename(filetypes=([('Config files', '.conf')]))
			if fn != '':
				global multiselect, select_symbol, select_color
				pre_select = multiselect
				pre_basemap = self.basemap_var.get()
				pre_symbol = select_symbol
				pre_color = select_color
				pre_loc_symbol = location_marker
				pre_loc_color = location_color
				pre_label_color = label_color
				pre_label_font = label_font
				pre_label_size = label_size
				pre_label_bold = label_bold
				pre_label_position = label_position
				read_config(fn)
				# (Re)set configuration options to global defaults
				self.map_option_menu['values'] = self.available_tile_servers()
				if multiselect != pre_select:
					self.multiselect_var.set(multiselect)
				if initial_basemap != pre_basemap:
					self.basemap_var.set(initial_basemap)
					tileserver = self.tile_url(initial_basemap)
					self.map_widget.set_tile_server(tileserver)
				if select_symbol != pre_symbol or select_color != pre_color:
					self.set_sel_marker(select_symbol, select_color)
				# Redraw markers if any setting has changed
				if location_marker != pre_loc_symbol or location_color != pre_loc_color or \
						label_color != pre_label_color or label_font != pre_label_font or \
						label_size != pre_label_size or label_bold != pre_label_bold or \
						label_position != pre_label_position:
							if label_font != pre_label_font or label_size != pre_label_size or label_bold != pre_label_bold:
								self.mapfont = self.makefont()
							self.loc_marker_icon = self.set_get_loc_marker()
							self.redraw_loc_markers(self.tbl)
							self.redraw_sel_markers()
				global config_files_user
				config_files_user.append(os.path.abspath(fn))
		def save_config():
			fn = tkfiledialog.asksaveasfilename(filetypes=([('Config files', '.conf')]))
			if fn != '':
				f = open(fn, "w")
				f.write("# Configuration file for mapdata.py\n# Created by export from mapdata.py at %s\n" % datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
				f.write("\n[basemap_tile_servers]\n")
				added_bms = [k for k in bm_servers if not k in initial_bms]
				for k in added_bms:
					f.write("%s=%s\n" % (k, bm_servers[k]))
				f.write("\n[api_keys]\n")
				for k in api_keys:
					f.write("%s=%s\n" % (k, api_keys[k]))
				f.write("\n[symbols]\n")
				for s in imported_symbols:
					f.write("%s=%s\n" % (s[0], s[1]))
				f.write("\n[defaults]\n")
				f.write("basemap=%s\n" % self.basemap_var.get())
				f.write("location_marker=%s\n" % location_marker)
				f.write("location_color=%s\n" % location_color)
				f.write("use_data_marker=%s\n" % use_data_marker)
				f.write("use_data_color=%s\n" % use_data_color)
				f.write("multilselect=%s\n" % ('Yes' if self.multiselect_var.get() == '1' else 'No'))
				f.write("select_symbol=%s\n" % select_symbol)
				f.write("select_color=%s\n" % select_color)
				f.write("label_color=%s\n" % label_color)
				f.write("label_font=%s\n" % label_font)
				f.write("label_size=%s\n" % label_size)
				f.write("label_bold=%s\n" % ('No' if not label_bold else 'Yes'))
				f.write("label_position=%s\n" % label_position)
				if editor is not None:
					f.write("\n[config]\n")
					f.write("editor=%s\n" % editor)

		def show_data_types():
			dlg = MsgDialog("Data Types", "Data types, data completeness, and number of unique\nnon-missing values for columns of the data table:")
			tframe, tdata = treeview_table(dlg.content_frame, self.data_types, ["Column", "Type", "Missing", "Unique"], "browse")
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			dlg.show()
		def run_query():
			dlg = QueryDialog(self.headers, self.db, self.whereclause)
			whereclause, action = dlg.get_where()
			if whereclause is not None:
				sqlcmd = "SELECT treeviewid FROM mapdata WHERE %s" % whereclause
				cur = self.db.cursor()
				try:
					result = cur.execute(sqlcmd)
					id_list = [r[0] for r in result.fetchall()]
				except:
					cur.close()
					warning("Invalid data selection expression: %s" % whereclause, kwargs={})
				else:
					cur.close()
					self.whereclause = whereclause
					# Enable multiselect
					global multiselect
					multiselect = "1"
					self.multiselect_var.set("1")
					self.tbl.configure(selectmode = tk.EXTENDED)
					if action == "Replace":
						# Remove any existing selections
						self.unselect_map()
						self.tbl.selection_set(tuple(id_list))
					elif action == "Add":
						# Add new selections
						all_selections = tuple(set(self.tbl.selection()) | set(id_list))
						self.tbl.selection_set(all_selections)
					else:
						# Remove
						diff_selections = tuple(set(self.tbl.selection()) - set(id_list))
						self.tbl.selection_set(diff_selections)
					self.mark_map(None)
					self.set_status()
		def invert_selections():
			selected = self.tbl.selection()
			new_selections = []
			for iid in self.tbl.get_children():
				if not iid in selected:
					new_selections.append(iid)
			self.tbl.selection_set(tuple(new_selections))
			self.mark_map(None)
			self.set_status()
		def new_plot():
			dlg = PlotDialog(self, self.data_types)
			self.plot_list.append(dlg)
			dlg.show

		def online_help():
			webbrowser.open("https://mapdata.osdn.io", new=2, autoraise=True)
		def show_config_files():
			if len(config_files) == 0 and len(config_files_user) == 0:
				msg = "No configuration files have been read."
			else:
				if len(config_files) > 0:
					msg = "Configuration files read on startup:\n   %s" % "\n   ".join(config_files)
					if len(config_files_user) > 0:
						msg = msg + "\n\n"
				if len(config_files_user) > 0:
					msg = msg + "Configuration files read after startup, in sequence:\n   %s" % "\n   ".join(config_files_user)
			dlg = MsgDialog("Config files", msg)
			dlg.show()
		def show_about():
			message="""
               mapdata.py

     version: %s, %s
Copyright %s, R Dreas Nielsen
         License: GNU GPL3""" % (version, vdate, copyright)
			dlg = MsgDialog("About", message)
			dlg.show()

		file_menu.add_command(label="Open CSV", command = self.new_data_file, underline=5)
		file_menu.add_command(label="Open spreadsheet", command = self.new_spreadsheet_file, underline=5)
		file_menu.add_command(label="Open database", command = self.new_db_table, underline=5)
		file_menu.add_command(label="Import symbol", command = import_symbol_file, underline=0)
		file_menu.add_command(label="Read config", command = read_config_file, underline=0)
		file_menu.add_command(label="Save config", command = save_config, underline=0)
		file_menu.add_command(label="Quit", command = self.cancel, underline=0)
		tbl_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		tbl_menu.add_command(label="Export selected", command = save_table, underline=1)
		tbl_menu.add_command(label="Data types", command = show_data_types, underline=5)
		map_menu.add_command(label="Change marker", command = change_marker, underline=7)
		map_menu.add_command(label="Change labeling", command = change_labeling, underline=7)
		map_menu.add_command(label="Zoom selected", command = self.zoom_selected, underline=5)
		map_menu.add_command(label="Zoom full", command = self.zoom_full, underline=5)
		map_menu.add_command(label="Center on selection", command = self.focus_map, underline=0)
		map_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		map_menu.add_command(label="Change CRS", command = self.change_crs, underline=1)
		map_menu.add_command(label="Export", command = save_map, underline=1)
		sel_menu.add_command(label="Invert", command = invert_selections, underline=0)
		sel_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		sel_menu.add_command(label="Set by query", command = run_query, underline=7)
		plot_menu.add_command(label="New", command = new_plot, underline=0)
		plot_menu.add_command(label="Close all", command = self.close_all_plots, underline=0)
		help_menu.add_command(label="Online help", command = online_help, underline=7)
		help_menu.add_command(label="Config files", command = show_config_files, underline=0)
		help_menu.add_command(label="About", command = show_about, underline=0)



class LabelDialog(object):
	def __init__(self, parent, column_list, label_col):
		columns = ['']
		columns.extend(column_list)
		label_col = '' if label_col is None else label_col
		self.dlg = tk.Toplevel()
		self.dlg.title("Change Labeling")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		#
		symbol_lbl = ttk.Label(prompt_frame, text="Location symbol:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.symbol_var = tk.StringVar(self.dlg, location_marker)
		symbol_vals = list(icon_xbm.keys())
		symbol_vals.sort()
		self.symbol_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.symbol_var,
				values=symbol_vals, width=15)
		self.symbol_opts.grid(row=0, column=1, columnspan=3, sticky=tk.W, padx=(6,3))
		color_lbl = ttk.Label(prompt_frame, text="Color:")
		color_lbl.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.color_var = tk.StringVar(self.dlg, location_color)
		color_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.color_var,
				values=list(select_colors), width=15)
		color_opts.grid(row=1, column=1, columnspan=3, sticky=tk.W, padx=(6,3), pady=(3,3))
		#
		self.use_data_marker_var = tk.StringVar(prompt_frame, use_data_marker)
		ck_use_data_marker = ttk.Checkbutton(prompt_frame, text="Use data symbol", variable=self.use_data_marker_var)
		ck_use_data_marker.grid(row=0, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.use_data_color_var = tk.StringVar(prompt_frame, use_data_color)
		ck_use_color_marker = ttk.Checkbutton(prompt_frame, text="Use data color", variable=self.use_data_color_var)
		ck_use_color_marker.grid(row=1, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.col_var = tk.StringVar(prompt_frame, label_col)
		col_lbl = ttk.Label(prompt_frame, text = "Data column:")
		col_lbl.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(6,3))
		col_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.col_var,
				values=columns, width=40)
		col_opts.grid(row=2, column=1, columnspan=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.font_var = tk.StringVar(prompt_frame, label_font)
		font_lbl = ttk.Label(prompt_frame, text="Font:")
		font_lbl.grid(row=3, column=0, sticky=tk.E, padx=(3,3), pady=(6,3))
		fonts = list(set(list(tkfont.families())))
		fonts.sort()
		font_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.font_var,
				values=fonts, width=25)
		font_opts.grid(row=3, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.bold_var = tk.StringVar(prompt_frame, "0" if not label_bold else "1")
		ck_bold = ttk.Checkbutton(prompt_frame, text="Bold", variable=self.bold_var)
		ck_bold.grid(row=3, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.size_var = tk.IntVar(prompt_frame, label_size)
		#
		size_lbl = ttk.Label(prompt_frame, text="Size:")
		size_lbl.grid(row=4, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		size_opt = ttk.Combobox(prompt_frame, state="normal", textvariable=self.size_var,
				values=[8, 10, 12, 14, 16, 20, 24], width=3)
		size_opt.grid(row=4, column=1, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.position_var = tk.StringVar(prompt_frame, label_position)
		position_lbl = ttk.Label(prompt_frame, text="Position:")
		position_lbl.grid(row=4, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		position_sel = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.position_var,
				values=["above", "below"], width=6)
		position_sel.grid(row=4, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#change-labeling", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_labeling(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_opts.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.color_var.get(), self.use_data_marker_var.get(),
					self.use_data_color_var.get(), self.col_var.get(), self.font_var.get(), self.size_var.get(),
					self.bold_var.get(), self.position_var.get())
		else:
			return (None,None,None,None,None,None,None,None,None)



class MarkerDialog(object):
	def __init__(self, parent):
		self.dlg = tk.Toplevel()
		self.dlg.title("Change Marker")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		symbol_lbl = ttk.Label(prompt_frame, text="Marker symbol:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.symbol_var = tk.StringVar(self.dlg, select_symbol)
		symbol_vals = list(icon_xbm.keys())
		symbol_vals.sort()
		self.symbol_opts = ttk.Combobox(prompt_frame, state="readonly", textvar=self.symbol_var,
				values=symbol_vals, width=15)
		self.symbol_opts.grid(row=0, column=1, sticky=tk.W, padx=(3,6))
		color_lbl = ttk.Label(prompt_frame, text="Color:")
		color_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		self.color_var = tk.StringVar(self.dlg, select_color)
		color_vals = list(select_colors)
		color_opts = ttk.Combobox(prompt_frame, state="readonly", textvar=self.color_var,
				values=color_vals, width=15)
		color_opts.grid(row=1, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#change-marker", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_marker(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_opts.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.color_var.get())
		else:
			return (None, None)


class ImportSymbolDialog(object):
	def __init__(self):
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('X11 bitmaps', '.xbm')]))
			if fn != '':
				self.fn_var.set(fn)
		def check_enable(*args):
			if self.fn_var.get() != '' and self.symbol_var.get() != '':
				self.ok_btn["state"] = tk.NORMAL
			else:
				self.ok_btn["state"] = tk.DISABLED
		self.dlg = tk.Toplevel()
		self.dlg.title("Import X11 Bitmap Symbol")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		button_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		button_frame.grid(row=1, column=0, columnspan=3, sticky=tk.EW, pady=(3,3))
		button_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(button_frame)
		btn_frame.grid(row=0, column=0, sticky=tk.EW)
		btn_frame.columnconfigure(0, weight=1)
		# Prompts
		symbol_lbl = ttk.Label(prompt_frame, text="Symbol name:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.symbol_var = tk.StringVar(self.dlg, "")
		self.symbol_var.trace('w', check_enable)
		self.symbol_entry = ttk.Entry(prompt_frame, textvariable=self.symbol_var, width=12)
		self.symbol_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		fn_label = ttk.Label(prompt_frame, text="File:")
		fn_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(prompt_frame, '')
		self.fn_var.trace('w', check_enable)
		fn_entry = ttk.Entry(prompt_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=1, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(prompt_frame, text="Browse", command=get_fn)
		fn_btn.grid(row=1, column=2, sticky=tk.W)
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn["state"] = tk.DISABLED
		self.ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#import-symbol", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.ok_btn["state"] != tk.DISABLED:
			self.canceled = False
			self.dlg.destroy()
	def run(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_entry.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.fn_var.get())
		else:
			return (None, None)



class DataFileDialog(object):
	def __init__(self):
		def get_fn():
			fn = tkfiledialog.askopenfilename(filetypes=([('CSV files', '.csv')]), parent=self.dlg)
			if fn != '':
				self.fn_var.set(fn)
				csvreader = CsvFile(fn)
				self.header_list = csvreader.next()
				self.id_sel["values"] = self.header_list
				self.lat_sel["values"] = self.header_list
				self.lon_sel["values"] = self.header_list
				self.sym_sel["values"] = self.header_list
				self.col_sel["values"] = self.header_list
		def check_enable(*args):
			if self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
				ok_btn["state"] = tk.NORMAL
			else:
				ok_btn["state"] = tk.DISABLED
		def new_fn(*args):
			check_enable()
		self.header_list = []
		self.dlg = tk.Toplevel()
		self.dlg.title("Open CSV Data File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		dir_frame = tk.Frame(prompt_frame)
		dir_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		dir_frame.rowconfigure(0, weight=1)
		dir_frame.columnconfigure(0, weight=1)
		req_frame = ttk.LabelFrame(prompt_frame, text="Required")
		req_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(3,3))
		req_frame.columnconfigure(0, weight=1)
		opt_frame = ttk.LabelFrame(prompt_frame, text="Optional")
		opt_frame.grid(row=2, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		opt_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Prompts
		#-- Directions
		dir_lbl = ttk.Label(dir_frame,
				text="Select a CSV file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(3,3), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)
		#-- Filename
		fn_frame = tk.Frame(req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(5,5))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		#-- Required columns
		column_choices = list(self.header_list)
		#
		req_col_frame = tk.Frame(req_frame)
		req_col_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		lat_label = ttk.Label(req_col_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(req_col_frame, '')
		self.lat_var.trace('w', check_enable)
		self.lat_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=12)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		#
		lon_label = ttk.Label(req_col_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(20,3), pady=(3,3))
		self.lon_var = tk.StringVar(req_frame, '')
		self.lon_var.trace('w', check_enable)
		self.lon_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=12)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		#-- Optional columns
		opt_col_frame = tk.Frame(opt_frame)
		opt_col_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		id_label = ttk.Label(opt_col_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(opt_col_frame, '')
		self.id_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		crs_label = ttk.Label(opt_col_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(opt_col_frame, 4326)
		self.crs_var.trace('w', check_enable)
		self.crs_sel = ttk.Entry(opt_col_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		sym_label = ttk.Label(opt_col_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(opt_col_frame, '')
		self.sym_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		col_label = ttk.Label(opt_col_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(opt_col_frame, '')
		self.col_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#-- Description
		title_label = ttk.Label(opt_col_frame, text="Description:")
		title_label.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.title_var = tk.StringVar(opt_col_frame, '')
		title_entry = ttk.Entry(opt_col_frame, width=60, textvariable=self.title_var)
		title_entry.grid(row=2, column=1, columnspan=3, sticky=tk.EW, padx=(3,6), pady=(3,3))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		ok_btn["state"] = tk.DISABLED
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.resizable(False, False)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#open-csv-data-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			headers, rows = file_data(self.fn_var.get())
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.title_var.get(),
					headers, rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)



class ImportSpreadsheetDialog(object):
	def __init__(self):
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('Spreadsheets', '.ods .xlsx .xls')]), parent=self.dlg)
			if fn != '':
				self.fn_var.set(fn)
		def check_w1enable(*args):
			if self.fn_var.get() != '':
				if os.path.isfile(self.fn_var.get()):
					w1next_btn["state"] = tk.NORMAL
				else:
					w1next_btn["state"] = tk.DISABLED
			else:
				w1next_btn["state"] = tk.DISABLED
		def check_w2enable(*args):
			if self.fn_var.get() != '' and self.sheet_var.get() != '':
				w2next_btn["state"] = tk.NORMAL
			else:
				w2next_btn["state"] = tk.DISABLED
		def check_w3enable(*args):
			if self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
				w3ok_btn["state"] = tk.NORMAL
			else:
				w3ok_btn["state"] = tk.DISABLED
		def new_fn(*args):
			check_w1enable()
		self.sheet_list = []
		self.header_list = []
		self.dlg = tk.Toplevel()
		self.dlg.title("Open Spreadsheet File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Wizard frames 1, 2, and 3 are gridded in the same cell to make a wizard.
		self.dlg.rowconfigure(0, weight=0)
		wiz1_frame = tk.Frame(self.dlg)		# For description, filename, and sheet name
		wiz1_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz1_frame.rowconfigure(0, weight=1)
		wiz2_frame = tk.Frame(self.dlg)		# For sheet selections
		wiz2_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame = tk.Frame(self.dlg)		# For column selections
		wiz3_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz3_frame.columnconfigure(0, weight=1)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.resizable(False, False)
		wiz1_frame.lift()

		# Populate directions frame
		dir_lbl = ttk.Label(prompt_frame,
				text="Select a spreadsheet file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)

		# Populate wiz1_frame
		w1req_frame = ttk.LabelFrame(wiz1_frame, text="Required")
		w1req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w1req_frame.columnconfigure(0, weight=1)
		fn_frame = tk.Frame(w1req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(3,3))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn, underline=0)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-b>", get_fn)

		w1opt_frame = ttk.LabelFrame(wiz1_frame, text="Optional")
		w1opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		w1opt_frame.columnconfigure(0, weight=1)
		desc_label = ttk.Label(w1opt_frame, text="Description:")
		desc_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.desc_var = tk.StringVar(w1opt_frame, '')
		desc_entry = ttk.Entry(w1opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w1_next(*args):
			if self.fn_var.get() != '':
				# Open spreadsheet, get sheet names
				fn, ext = os.path.splitext(self.fn_var.get())
				ext = ext.lower()
				try:
					if ext == '.ods':
						sso = OdsFile()
					elif ext == '.xlsx':
						sso = XlsxFile()
					else:
						sso = XlsFile()
				except:
					warning("Could not open %s" % self.fn_var.get(), kwargs={'parent': self.dlg})
				else:
					sso.open(self.fn_var.get())
					self.sheet_list = sso.sheetnames()
					self.sheet_sel["values"] = self.sheet_list
					if ext in ('.ods', '.xlsx'):
						try:
							sso.close()
						except:
							pass
					else:
						try:
							sso.release_resources()
							del sso
						except:
							pass
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz2_frame.lift()
					self.dlg.bind("<Alt-b>", w2_back)
					self.dlg.bind("<Alt-n>", w2_next)

		w1btn_frame = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)
		w1btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w1btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		#
		w1help_btn = ttk.Button(w1btn_frame, text="Help", command=self.do_help, underline=0)
		w1help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = ttk.Button(w1btn_frame, text="Next", command=w1_next, underline=0)
		w1next_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		self.dlg.bind("<Alt-n>", w1_next)
		w1cancel_btn = ttk.Button(w1btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w1cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		w1next_btn["state"] = tk.DISABLED
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)


		# Populate wiz2_frame
		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(0, weight=1)
		w2inner_frame = tk.Frame(w2req_frame)
		w2inner_frame.grid(row=0, column=0, sticky=tk.W)
		#
		sheet_label = ttk.Label(w2inner_frame, text="Sheet:")
		sheet_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.sheet_var = tk.StringVar(w2req_frame, '')
		self.sheet_var.trace('w', check_w2enable)
		self.sheet_sel = ttk.Combobox(w2inner_frame, state="readonly", textvariable=self.sheet_var, values=self.sheet_list, width=16)
		self.sheet_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		xrows_label = ttk.Label(w2inner_frame, text="Initial rows to skip:")
		xrows_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.xrows_var = tk.IntVar(w2req_frame, 0)
		self.xrows_var.trace('w', check_w2enable)
		xrows_entry = ttk.Entry(w2inner_frame, textvariable=self.xrows_var, width=6)
		xrows_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,3))

		def w2_back(*args):
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-n>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>", get_fn)

		def w2_next(*args):
			# Open spreadsheet, get column names
			if self.fn_var.get() != '' and self.sheet_var.get() != '':
				fn, ext = os.path.splitext(self.fn_var.get())
				try:
					if ext.lower() == '.ods':
						hdrs, data = ods_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
					else:
						hdrs, data = xls_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
				except:
					warning("Could not read table from %s, sheet %s" % (self.fn_var.get(), self.sheet_var.get()), 
							kwargs={'parent': self.dlg})
				else:
					self.headers = hdrs
					self.header_list = list(hdrs)
					self.rows = data
					# Set list box values
					self.id_sel["values"] = self.header_list
					self.lat_sel["values"] = self.header_list
					self.lon_sel["values"] = self.header_list
					self.sym_sel["values"] = self.header_list
					self.col_sel["values"] = self.header_list
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz3_frame.lift()
					self.dlg.bind("<Alt-b>", w3_back)

		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = ttk.Button(w2btn_frame, text="Help", command=self.do_help, underline=0)
		w2help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w2prev_btn = ttk.Button(w2btn_frame, text="Back", command=w2_back, underline=0)
		w2prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w2next_btn = ttk.Button(w2btn_frame, text="Next", command=w2_next, underline=0)
		w2next_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w2cancel_btn = ttk.Button(w2btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w2cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w2next_btn["state"] = tk.DISABLED
	
		# Populate wiz3_frame
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=15)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=15)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))


		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(9,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w2enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w3_back(*args):
			self.dlg.bind("<Alt-b>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-b>", w2_back)
			self.dlg.bind("<Alt-n>", w2_next)

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = ttk.Button(w3btn_frame, text="Help", command=self.do_help, underline=0)
		w3help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w3prev_btn = ttk.Button(w3btn_frame, text="Back", command=w3_back)
		w3prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w3ok_btn = ttk.Button(w3btn_frame, text="OK", command=self.do_select, underline=0)
		w3ok_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		self.dlg.bind('<Alt-o>', self.do_select)
		w3cancel_btn = ttk.Button(w3btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w3cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w3ok_btn["state"] = tk.DISABLED
	
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#open-spreadsheet-data-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(),
					self.headers, self.rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)


class DbConnectDialog(object):
	FILE, SERVER, FILE_PW = range(3)
	def __init__(self):
		self.exit_status = 0	# Canceled
		self.exit_svr = None	# For caller
		self.exit_db = None	# For caller
		self.xpos = None
		self.ypos = None
		self.scriptfilepath = None
		# Values of db_params indicate whether server information is needed.
		self.db_params = {u"PostgreSQL": self.SERVER, u"SQLite": self.FILE, u"DuckDB": self.FILE,
							u"SQL Server": self.SERVER, u"MySQL": self.SERVER, u"Firebird": self.SERVER,
							u"MariaDB": self.SERVER, u"Oracle": self.SERVER}
		self.dlg = tk.Toplevel()
		self.title = "Database Table for Map Display"
		self.dlg.title(self.title)
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.dlg.geometry("650x260")
		self.headers = None
		self.header_list = None
		self.datarows = None

		# Main frames
		msgframe = tk.Frame(self.dlg)
		msgframe.grid(column=0, row=0, padx=6, pady=2, sticky=tk.EW)
		# Database selection is in one wizard pane, table and script selection are in a second, and column selection is in a third wizard pane
		wiz1_frame = tk.Frame(self.dlg)
		wiz1_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz2_frame = tk.Frame(self.dlg)
		wiz2_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz3_frame = tk.Frame(self.dlg)
		wiz3_frame.grid(column=0, row=1, sticky=tk.NSEW)
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		wiz1_frame.rowconfigure(0, weight=1)
		wiz1_frame.columnconfigure(0, weight=1)
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame.rowconfigure(0, weight=1)
		wiz3_frame.columnconfigure(0, weight=1)

		# Populate message frame
		msg_label = ttk.Label(msgframe, text="The database, table, and columns to be used for mapping must be specified.", anchor=tk.W, justify=tk.LEFT, wraplength=500)
		msg_label.grid(column=0, row=0, sticky=tk.EW)

		# Wizard page 1
		# Database selector
		# On the left side will be a combobox to choose the database type.
		# On the right side will be a prompt for the server, db, user name, and pw,
		# or for the filename (and possibly user name and pw).  Each of these alternative
		# types of prompts will be in its own frame, which will be in the same place.
		# Only one will be shown, controlled by the item in the self.db_params dictionary.
		# A separate frame for the table name is below the database parameters frame.
		dbframe = tk.Frame(wiz1_frame)
		dbtypeframe = tk.Frame(dbframe)
		rightframe = tk.Frame(dbframe)
		paramframe = tk.Frame(rightframe)
		self.serverparamframe = tk.Frame(paramframe)
		self.fileparamframe = tk.Frame(paramframe)
		self.filepwparamframe = tk.Frame(paramframe)
		w1btnframe = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)

		# Grid wiz1 frame widgets
		def param_choices(*args, **kwargs):
			svr_params = self.db_params[self.db_type_var.get()]
			if svr_params == self.SERVER:
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.serverparamframe.grid()
			elif svr_params == self.FILE_PW:
				self.serverparamframe.grid_remove()
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid()
			else:
				self.serverparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.fileparamframe.grid()
			check_w1enable()

		def check_w1enable(*args):
			dbms = self.db_type_var.get()
			if dbms != '':
				dbtype = self.db_params[dbms]
				if dbtype == self.SERVER:
					if self.server.get() != '' and self.db.get != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				elif dbtype == self.FILE_PW:
					if self.db_file.get() != '' and self.user.get() != '' and self.pw.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				else:	# self.FILE
					if self.db_file.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED

		dbframe.grid(column=0, row=0, sticky=tk.NSEW)
		dbtypeframe.grid(column=0, row=0, padx=5, sticky=tk.NW)
		rightframe.grid(column=1, row=0, padx=5, sticky=tk.N + tk.EW)
		paramframe.grid(column=0, row=0, padx=5, sticky=tk.N + tk.EW)
		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe.
		# Leave only serverparamframe visible.
		self.fileparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		w1btnframe.grid(column=0, row=2, sticky=tk.S+tk.EW)
		w1btnframe.columnconfigure(0, weight=1)


		# Populate dbframe
		self.db_type_var = tk.StringVar()
		self.encoding = tk.StringVar()
		self.table_var = tk.StringVar()
		self.table_var.trace('w', check_w1enable)
		# Database type selection
		ttk.Label(dbtypeframe, text="DBMS:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NE)
		dbmss = [k for k in self.db_params.keys()]
		dbmss.sort()
		self.db_choices = ttk.Combobox(dbtypeframe, textvariable=self.db_type_var, width=12,
						values=dbmss)
		self.db_choices.bind("<<ComboboxSelected>>", param_choices)
		self.db_choices.config(state='readonly')
		self.db_choices.grid(column=1, row=0, padx=3, pady=3, sticky=tk.NW)
		ttk.Label(dbtypeframe, text="Encoding:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.NE)
		self.db_choices.set('PostgreSQL')
		enc_choices = ttk.Combobox(dbtypeframe, textvariable=self.encoding, width=12,
						values=('UTF8', 'Latin1', 'Win1252'))
		enc_choices.set('UTF8')
		enc_choices.grid(column=1, row=1, padx=3, pady=3, sticky=tk.NW)
		# Database parameter entry frames
		self.server = tk.StringVar()
		self.server.trace('w', check_w1enable)
		self.port = tk.StringVar()
		self.db = tk.StringVar()
		self.db.trace('w', check_w1enable)
		self.user = tk.StringVar()
		self.user.trace('w', check_w1enable)
		self.pw = tk.StringVar()
		self.pw.trace('w', check_w1enable)
		self.db_file = tk.StringVar()
		self.db_file.trace('w', check_w1enable)

		# Server databases
		ttk.Label(self.serverparamframe, text="Server:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.server).grid(column=1, row=0, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Database:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.db).grid(column=1, row=1, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Port:").grid(column=0, row=4, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=4, textvariable=self.port).grid(column=1, row=4, padx=3, pady=3, sticky=tk.W)

		# File databases
		ttk.Label(self.fileparamframe, text="Database file:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.fileparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.fileparamframe, text="Browse...", command=self.set_sel_fn).grid(column=1, row=1)

		# File databases with user name and password
		ttk.Label(self.filepwparamframe, text="Database file:").grid(column=0, row=0, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.filepwparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.filepwparamframe, text="Browse...", command=self.set_sel_fn).grid(column=2, row=1)
		ttk.Label(self.filepwparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.filepwparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)

		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe
		self.fileparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.db_type_var.set(u"PostgreSQL")

		def w1_next(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			wiz2_frame.lift()
			# The following conditional fails
			#if w1next_btn["state"] == tk.NORMAL:
			#	wiz2_frame.lift()
			self.dlg.bind("<Alt-n>")
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)

		# Populate w1btnframe
		w1help_btn = ttk.Button(w1btnframe, text="Help", command=self.do_help, underline=0)
		w1help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = ttk.Button(w1btnframe, text="Next", command=w1_next, underline=0)
		w1next_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w1cancel_btn = ttk.Button(w1btnframe, text="Cancel", command=self.do_cancel, underline=0)
		w1cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		w1next_btn["state"] = tk.DISABLED
		self.dlg.bind("<Alt-n>", w1_next)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)
		

		# Wizard page 2
		# Database table and optional query
		def check_w2enable(*args):
			if self.table_var.get() != '':
				w2next_btn["state"] = tk.NORMAL
			else:
				w2next_btn["state"] = tk.DISABLED
		def w2_back(*args):
			self.dlg.unbind("<Alt-p>")
			self.dlg.unbind("<Alt-s>")
			self.dlg.unbind("<Alt-e>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>")
		def w2_next(*args):
			if self.table_var.get() != '':
				sql = "select * from %s;" % self.table_var.get()
				conn = None
				# Open database, get table data and column headers, populate wiz2 comboboxes
				dbms = self.db_type_var.get()
				if dbms == 'PostgreSQL':
					try:
						import psycopg2
					except:
						warning("The Python package 'psycopg2' must be installed.", kwargs={'parent': self.dlg})
					else:
						try:
							port = '5432' if self.port.get() == '' else self.port.get()
							if self.user.get() != '' and self.pw != '':
								conn = psycopg2.connect(host=self.server.get(), database=self.db.get(), port=port, user=self.user.get(), password=self.pw.get())
							else:
								conn = psycopg2.connect(host=self.server.get(), database=self.db.get(), port=port)
						except:
							warning("Cannot open the Postgres database.", kwargs={'parent': self.dlg})

				elif dbms == 'SQLite':
					import sqlite3
					try:
						conn = sqlite3.connect(self.db_file.get())
					except:
						warning("Cannot open the file %s as a SQLite database." % self.db_file.get(), kwargs={'parent': self.dlg})

				elif dbms == 'DuckDB':
					try:
						import duckdb
					except:
						warning("The Python package 'duckdb' must be installed.", kwargs={'parent': self.dlg})
					else:
						try:
							conn = duckdb.connect(self.db_file.get(), read_only=True)
						except:
							warning("Cannot open the file %s as a DuckDB database." % self.db_file.get(), kwargs={'parent': self.dlg})

				elif dbms == 'MariaDB' or dbms == 'MySQL':
					try:
						import pymysql
					except:
						warning("The Python package 'pymysql' must be installed.", kwargs={'parent': self.dlg})
					else:
						try:
							port = '3306' if self.port.get() == '' else self.port.get()
							if self.user.get() != '' and self.pw != '':
								conn = pymysql.connect(host=self.server.get(), database=self.db.get(), port=port, user=self.user.get(), password=self.pw.get())
							else:
								conn = pymysql.connect(host=self.server.get(), database=self.db.get(), port=port)
						except:
							warning("Cannot open the MariaDB/MySQL database.", kwargs={'parent': self.dlg})

				elif dbms == 'SQL Server':
					try:
						import pyodbc
					except:
						warning("The Python package 'pyodbc' must be installed.", kwargs={'parent': self.dlg})
					else:
						ssdrivers = ('ODBC Driver 17 for SQL Server', 'ODBC Driver 13.1 for SQL Server',
							'ODBC Driver 13 for SQL Server', 'ODBC Driver 11 for SQL Server',
							'SQL Server Native Client 11.0', 'SQL Server Native Client 10.0',
							'SQL Native Client', 'SQL Server')
						for drv in ssdrivers:
							if self.user.get() != '':
								if self.pw.get() != '':
									connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s;Pwd=%s" % (drv, self.server.get(), self.db.get(), self.user.get(), self.pw.get())
								else:
									connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s;Pwd=%s" % (drv, self.server.get(), self.db.get(), self.user.get())
							else:
								connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Trusted_Connection=yes" % (drv, self.server.get(), self.db.get())
							try:
								conn = pyodbc.connect(connstr)
							except:
								conn = None
							else:
								break
						if conn == None:
							warning("Cannot open the SQL Server database.", kwargs={'parent': self.dlg})

				elif dbms == 'Oracle':
					port = '1521' if self.port.get() == '' else self.port.get()
					try:
						import cx_Oracle as orc
					except:
						warning("The Python package 'cx-Oracle' must be installed.", kwargs={'parent': self.dlg})
					else:
						try:
							dsn = orc.makedsn(self.server.get(), port, service_name=self.db.get())
							if self.user.get() != '' and self.pw != '':
								conn = orc.connect(user=self.user.get(), password=self.pw.get(), dsn=dsn)
							else:
								conn = orc.connect(dsn=dsn)
						except:
							warning("Cannot open the Oracle database.", kwargs={'parent': self.dlg})

				elif dbms == 'Firebird':
					port = '3050' if self.port.get == '' else self.port.get()
					try:
						import fdb
					except:
						warning("The Python package 'fdb' must be installed.", kwargs={'parent': self.dlg})
					else:
						try:
							if self.user.get() != '' and self.pw != '':
								return fdb.connect(host=self.server.get(), database=self.db.get(), port=port, user=self.user.get(), password=self.pw.get())
							else:
								return fdb.connect(host=self.server.get(), database=self.db.get(), port=port)
						except:
							warning("Cannot open the Firebird database.", kwargs={'parent': self.dlg})

				else:
					warning("Unrecognized DBMS type", kwargs={'parent': self.dlg})

				if conn is not None:
					curs = conn.cursor()
					sqlscript = self.script_text.get("1.0", "end-1c")
					run_ok = True
					if sqlscript != '':
						try:
							curs.execute(sqlscript)
						except:
							run_ok = False
							warning("Execution of the SQL script failed.", kwargs={'parent':self.dlg})
					if run_ok:
						try:
							curs.execute(sql)
						except:
							warning("Cannot select data from table %s." % self.table_var.get(), kwargs={'parent':self.dlg})
						else:
							self.headers = [d[0] for d in curs.description]
							self.rows = curs.fetchall()
							self.header_list = list(self.headers)
							# Set list box values
							self.id_sel["values"] = self.header_list
							self.lat_sel["values"] = self.header_list
							self.lon_sel["values"] = self.header_list
							self.sym_sel["values"] = self.header_list
							self.col_sel["values"] = self.header_list
							self.dlg.unbind("<Alt-p>")
							self.dlg.unbind("<Alt-s>")
							self.dlg.unbind("Alt-e>")
							wiz3_frame.lift()
							self.dlg.bind("<Alt-n>")
							self.dlg.bind("<Alt-b>")
							self.dlg.bind("<Alt-b>", w3_back)
							self.dlg.bind('<Alt-o>', self.do_select)
					conn.close()

		def load_script(*args):
			fn = tkfiledialog.askopenfilename(parent=self.dlg, title="SQL script file to open", filetypes=([('SQL script files', '.sql')]))
			if not (fn is None or fn == ''):
				path, filename = os.path.split(os.path.abspath(fn))
				self.scriptfilepath = path
				with open(fn, "r") as f:
					sql = f.read()
				self.script_text.insert("end", sql)
		def save_script(*args):
			outfile = tkfiledialog.asksaveasfilename(initialdir=self.scriptfilepath, parent=self.dlg, title="SQL script file to save", filetypes=[('SQL script files', '.sql')])
			if not (outfile is None or outfile == ''):
				sql = self.script_text.get("1.0", "end")
				with open(outfile, "w") as f:
					f.write(sql)
		def edit_sql(*args):
			td = tempfile.TemporaryDirectory()
			edit_fn = os.path.join(td.name, "mapfile_temp.sql")
			with open(edit_fn, "w") as f:
				f.write(self.script_text.get("1.0", "end"))
			returncode = subprocess.call([editor, edit_fn])
			if returncode == 0:
				with open(edit_fn, "r") as f:
					sql = f.read()
				self.script_text.delete("1.0", "end")
				self.script_text.insert("end", sql)
			else:
				warning("Failure attempting to edit the SQL with %s" % editor)

		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(1, weight=1)
		#
		tbl_label = ttk.Label(w2req_frame, text="Table:")
		tbl_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.table_var = tk.StringVar(w2req_frame, '')
		self.table_var.trace('w', check_w2enable)
		ttk.Entry(w2req_frame, width=30, textvariable=self.table_var).grid(row=0, column=1, padx=(3,6), pady=3, sticky=tk.W)
		#
		w2opt_frame = ttk.LabelFrame(wiz2_frame, text="Optional")
		w2opt_frame.grid(row=1, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2opt_frame.columnconfigure(1, weight=1)
		#
		self.script_text = tk.Text(w2opt_frame, width=40, height=4)
		self.script_text.grid(row=0, column=1, columnspan=2, rowspan=4, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		scr_label = ttk.Label(w2opt_frame, text="Script:")
		scr_label.grid(row=0, column=0, sticky=tk.NE, padx=(6,3), pady=(2,2))
		load_btn = ttk.Button(w2opt_frame, text="Open", command=load_script, underline=1)
		load_btn.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		save_btn = ttk.Button(w2opt_frame, text="Save", command=save_script, underline=0)
		save_btn.grid(row=2, column=0, sticky=tk.E, padx=(3,3))
		edit_btn = ttk.Button(w2opt_frame, text="Edit", command=edit_sql, underline=0)
		edit_btn.grid(row=3, column=0, sticky=tk.E, padx=(3,3), pady=(0,2))
		if editor is None:
			edit_btn["state"] = tk.DISABLED
		else:
			edit_btn["state"] = tk.NORMAL
		sbar = tk.Scrollbar(w2opt_frame)
		sbar.grid(row=0, column=2, rowspan=4, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.script_text.yview)
		self.script_text.config(yscrollcommand = sbar.set)
		#
		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(2,2))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = ttk.Button(w2btn_frame, text="Help", command=self.do_help, underline=0)
		w2help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		w2prev_btn = ttk.Button(w2btn_frame, text="Back", command=w2_back, underline=0)
		w2prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w2next_btn = ttk.Button(w2btn_frame, text="Next", command=w2_next, underline=0)
		w2next_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w2cancel_btn = ttk.Button(w2btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w2cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w2next_btn["state"] = tk.DISABLED


		# Wizard page 3
		# Column selectors
		def check_w3enable(*args):
			if self.lat_var.get() != '' and self.lon_var.get() != '':
				w3ok_btn["state"] = tk.NORMAL
			else:
				w3ok_btn["state"] = tk.DISABLED
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=15)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=15)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(6,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w3enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		desc_label = ttk.Label(w3opt_frame, text="Description:")
		desc_label.grid(row=2, column=0, sticky=tk.E, padx=(3,3), pady=(3,6))
		self.desc_var = tk.StringVar(w3opt_frame, '')
		desc_entry = ttk.Entry(w3opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=2, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,6))

		def w3_back(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			self.dlg.bind("<Alt-o>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = ttk.Button(w3btn_frame, text="Help", command=self.do_help, underline=0)
		w3help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w3prev_btn = ttk.Button(w3btn_frame, text="Back", command=w3_back, underline=0)
		w3prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w3ok_btn = ttk.Button(w3btn_frame, text="OK", command=self.do_select, underline=0)
		w3ok_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w3cancel_btn = ttk.Button(w3btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w3cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w3ok_btn["state"] = tk.DISABLED

		wiz1_frame.lift()

		self.canceled = True
		# Limit resizing
		self.dlg.resizable(False, False)
		center_window(self.dlg)
	def set_sel_fn(self):
		fn = tkfiledialog.askopenfilename(parent=self.fileparamframe, title=self.title)
		if fn:
			self.db_file.set(fn)
			#self.clearstatus()
	def do_select(self, *args):
		if self.table_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#open-database-data-source", new=2, autoraise=True)
	def get_data(self):
		self.dlg.grab_set()
		self.dlg.focus_force()
		self.db_choices.focus()
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return (None, None, None, None, None, None, None, None, None, None)
		else:
			return self.table_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(), \
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(), \
					self.headers, self.rows





class NewCrsDialog(object):
	def __init__(self, current_crs):
		self.dlg = tk.Toplevel()
		self.dlg.title("Change CRS")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.rowconfigure(0, weight=1)
		crs_lbl = ttk.Label(prompt_frame, text="New CRS:")
		crs_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.crs_var = tk.IntVar(self.dlg, current_crs)
		self.crs_entry = ttk.Entry(prompt_frame, width=12, textvariable=self.crs_var)
		self.crs_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#change-crs", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_crs(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.crs_entry.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return self.crs_var.get()
		else:
			return None


class QueryDialog(object):
	def __init__(self, column_headers, db_conn, init_sql=""):
		self.dlg = tk.Toplevel()
		self.dlg.title("Query Data")
		self.canceled = True
		self.dlg.columnconfigure(0, weight=1)
		self.dlg.rowconfigure(1, weight=1)
		# Frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		query_frame = tk.Frame(self.dlg)
		query_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		query_frame.rowconfigure(0, weight=1)
		query_frame.columnconfigure(0, weight=1)
		query_frame.columnconfigure(1, weight=1)
		sql_frame = tk.Frame(query_frame)
		sql_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		sql_frame.rowconfigure(0, weight=1)
		sql_frame.columnconfigure(0, weight=1)
		col_frame = tk.Frame(query_frame)
		col_frame.grid(row=0, column=1, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=1)
		act_frame = tk.Frame(query_frame)
		act_frame.grid(row=1, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Prompt
		prompt_lbl = ttk.Label(prompt_frame, wraplength=300, justify=tk.LEFT, text="Enter an expression below to identify the data rows that you want to select.  The syntax of this expression should correspond to a SQL 'WHERE' clause.  Column names with non-alphanumeric characters should be double-quoted.  String literals should be single-quoted.  The '%' character is a wildcard.  Ctrl-Enter completes entry.")
		prompt_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(3,3))
		def wrap_prompt(event):
			prompt_lbl.configure(wraplength=event.width - 5)
		prompt_lbl.bind("<Configure>", wrap_prompt)
		# SQL entry
		self.sql = ""
		self.sql_text = tk.Text(sql_frame, width=60, height=8)
		if init_sql != "":
			self.sql_text.insert("1.0", init_sql)
		self.sql_text.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		sbar = tk.Scrollbar(sql_frame)
		sbar.grid(row=0, column=1, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.sql_text.yview)
		self.sql_text.config(yscrollcommand = sbar.set)
		# Column values
		col_lbl = ttk.Label(col_frame, text="Column values:")
		col_lbl.grid(row=0, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		col_var = tk.StringVar(col_frame, "")
		colsel = ttk.Combobox(col_frame, state="readonly", textvariable=col_var, values=column_headers, width=20)
		colsel.grid(row=1, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		tv_frame = tk.Frame(col_frame)
		tv_frame.grid(row=2, column=0, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=0)
		col_frame.rowconfigure(1, weight=0)
		col_frame.rowconfigure(2, weight=1)
		def colval_to_sql(event):
			item_iid = self.tv_tbl.identify('item', event.x, event.y)
			item_val = self.tv_tbl.item(item_iid, "values")[0]
			if not isfloat(item_val):
				item_val = "'%s'" % item_val
			self.sql_text.insert(tk.END, " "+item_val)
		def list_col_vals(event):
			curs = db_conn.cursor()
			colname = dquote(col_var.get())
			res = curs.execute('SELECT DISTINCT %s FROM mapdata ORDER BY %s' % (colname, colname))
			rowset = res.fetchall()
			for widget in tv_frame.winfo_children():
				widget.destroy()
			tblframe, self.tv_tbl = treeview_table(tv_frame, rowset, [col_var.get()])
			tblframe.grid(column=0, row=0, sticky=tk.NSEW)
			curs.close()
			self.tv_tbl.bind("<Double-1>", colval_to_sql)
		colsel.bind("<<ComboboxSelected>>", list_col_vals)
		# Action selection
		self.act_var = tk.StringVar(act_frame, "Replace")
		act_lbl = ttk.Label(act_frame, text="Action:")
		act_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		act_sel = ttk.Combobox(act_frame, state="readonly", textvariable=self.act_var, values=["Replace", "Add", "Remove"], width=8)
		act_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Control-Return>", self.do_select)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#query-data", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.sql = self.sql_text.get("1.0", "end-1c")
		self.dlg.destroy()
	def get_where(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.sql_text.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.sql, self.act_var.get())
		else:
			return (None, None)


class PlotDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		self.dlg = tk.Toplevel()
		self.dlg.title("Plot")
		self.dlg.columnconfigure(0, weight=1)
		self.auto_update = True

		def set_autoupdate():
			if self.autoupdate_var.get() == "1":
				self.auto_update = True
				self.q_redraw()
			else:
				self.auto_update = False

		# Message
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=70, text="Select the type of plot, columns for X and Y data, and whether to show all data or only selected data.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Controls
		ctrl_frame = tk.Frame(self.dlg)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)

		self.type_var = tk.StringVar(ctrl_frame, "")
		type_lbl = ttk.Label(ctrl_frame, text="Plot type:")
		type_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.type_sel = ttk.Combobox(ctrl_frame, state="readonly", textvariable=self.type_var, width=20,
				values=["Box plot", "Category counts", "Empirical CDF", "Histogram", "Line plot", "Scatter plot", "Y range plot"])
		self.type_sel.grid(row=0, column=1, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.type_sel.bind("<<ComboboxSelected>>", self.set_xy)

		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_redraw, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=(6,3), pady=(3,3))

		self.autoupdate_var = tk.StringVar(ctrl_frame, "1")
		self.autoupdate_ck = ttk.Checkbutton(ctrl_frame, text="Auto-update", command=set_autoupdate, variable=self.autoupdate_var,
				onvalue="1", offvalue="0")
		self.autoupdate_ck.grid(row=1, column=2, sticky=tk.W, padx=(3,3), pady=(3,3))

		self.x_var = tk.StringVar(ctrl_frame, "")
		x_lbl = ttk.Label(ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(ctrl_frame, state="disabled", textvariable=self.x_var, width=20)
		self.x_sel.grid(row=0, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.q_redraw)

		self.y_var = tk.StringVar(ctrl_frame, "")
		y_lbl = ttk.Label(ctrl_frame, text="Y column:")
		y_lbl.grid(row=1, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.y_sel = ttk.Combobox(ctrl_frame, state="disabled", textvariable=self.y_var, width=20)
		self.y_sel.grid(row=1, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.y_sel.bind("<<ComboboxSelected>>", self.q_redraw)

		self.xlog_var = tk.StringVar(ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(ctrl_frame, text="Log X", state="disabled", command=self.q_redraw, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.ylog_var = tk.StringVar(ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(ctrl_frame, text="Log Y", state="disabled", command=self.q_redraw, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=1, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		# Plot
		self.content_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		self.plotfig = Figure(dpi=100)
		self.plotfig.set_figheight(5)
		self.plotfig_canvas = FigureCanvasTkAgg(self.plotfig, self.content_frame)
		self.plot_nav = NavigationToolbar2Tk(self.plotfig_canvas, self.content_frame)
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=0)
		btn_frame.columnconfigure(2, weight=0)
		btn_frame.columnconfigure(3, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		self.plot_data_btn = ttk.Button(btn_frame, text="Plot Data", state="disabled", command=self.show_plot_data, underline=0)
		self.plot_data_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,6))
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=3, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/dialogs.html#plot-dialog", new=2, autoraise=True)

	def show_data(self, *args):
		# Show data that have been collected for plotting, but not summarized as needed for a particular plot type.
		if self.dataset is not None:
			dlg = MsgDialog("Source Data", "Original data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			dlg.show()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			dlg = MsgDialog("Data for Plotting", "Data to be plotted:")
			variables = len(self.plot_data)
			rowwise_data = []
			max_data_len = max([len(self.plot_data[i]) for i in range(variables)])
			for i in range(max_data_len):
				row = []
				for j in range(variables):
					try:
						# Boxplot data are not necessarily a full matrix
						row.append(self.plot_data[j][i])
					except:
						row.append(None)
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.plot_data_labels)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			dlg.show()

	def set_xy(self, *args):
		# Enable X and Y value selection, and set Combobox values based on plot type and column types.
		self.plotfig.clear()
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.draw()
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		self.data_btn["state"] = "disabled"
		self.plot_data_btn["state"] = "disabled"
		categ_columns = [c[0] for c in self.column_specs if c[1] == "string"]
		# quant_columns includes date and timestamp columns
		quant_columns = [c[0] for c in self.column_specs if c[1] != "string"]
		numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		date_columns = [c[0] for c in self.column_specs if c[1] == "date"]
		plot_type = self.type_var.get()
		self.x_var.set('')
		self.y_var.set('')
		if plot_type == "Category counts":
			self.x_sel["values"] = categ_columns
			self.xlog_ck["state"] = "disabled"
			self.x_sel["state"] = "readonly"
			self.y_sel["state"] = "disabled"
			self.ylog_ck["state"] = "disabled"
		elif plot_type in ("Histogram", "Empirical CDF"):
			self.x_sel["values"] = numeric_columns
			self.xlog_ck["state"] = "normal"
			self.x_sel["state"] = "readonly"
			self.y_sel["state"] = "disabled"
			self.ylog_ck["state"] = "disabled"
		elif plot_type == "Box plot":
			self.x_sel["values"] = list(set(categ_columns) | set(date_columns))
			self.xlog_ck["state"] = "disabled"
			self.y_sel["values"] = quant_columns
			self.x_sel["state"] = "readonly"
			self.y_sel["state"] = "readonly"
			self.ylog_ck["state"] = "normal"
		else:
			self.x_sel["values"] = quant_columns
			self.xlog_ck["state"] = "normal"
			self.y_sel["values"] = quant_columns
			self.x_sel["state"] = "readonly"
			self.y_sel["state"] = "readonly"
			self.ylog_ck["state"] = "normal"

	def q_redraw(self, *args):
		# Conditionally (re)draw the plot.
		plot_type = self.type_var.get()
		can_redraw = (plot_type in ("Category counts", "Empirical CDF", "Histogram") and self.x_var.get() != '') \
				or (plot_type in ("Scatter plot", "Line plot", "Box plot", "Y range plot") and self.x_var.get() != '' and self.y_var.get() != '')
		if can_redraw:
			self.plotfig.clear()
			self.plot_axes = self.plotfig.add_subplot(111)
			self.plotfig_canvas.draw()
			self.get_data()
			if self.dataset is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = "disabled"
		self.plot_data_btn["state"] = "disabled"
		self.dataset = None
		plot_type = self.type_var.get()
		column_list = [self.x_var.get()]
		if self.y_var.get() != '':
			column_list.append(self.y_var.get())
		if self.sel_only_var.get() == "1":
			dataset = self.parent.get_sel_data(column_list)
		else:
			dataset = self.parent.get_all_data(column_list)
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.plot_data = None
			self.plot_data_labels = None
			self.data_btn["state"] = "disabled"
			self.plot_data_btn["state"] = "disabled"
		else:
			# Remove missing data
			column_indexes = range(len(dataset))
			clean_data = [[] for _ in dataset]
			for i in range(len(dataset[0])):
				ok = True
				for col in column_indexes:
					if dataset[col][i] is None or dataset[col][i] == '':
						ok = False
				if ok:
					for col in column_indexes:
						clean_data[col].append(dataset[col][i])
			dataset = None
			# Convert quantitative data types
			if plot_type != "Category counts":
				x_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.x_var.get()][0]
				cast_fn = data_type_cast_fn(x_data_type)
				for i in range(len(clean_data[0])):
					clean_data[0][i] = cast_fn(clean_data[0][i])
			if self.y_sel["state"] != "disabled" and self.y_var.get() != "" and len(clean_data) > 1:
				y_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.y_var.get()][0]
				cast_fn = data_type_cast_fn(y_data_type)
				for i in range(len(clean_data[1])):
					clean_data[1][i] = cast_fn(clean_data[1][i])
			# Set data labels
			if self.y_var.get() != '':
				self.data_labels = [self.x_var.get(), self.y_var.get()]
			else:
				self.data_labels = [self.x_var.get()]
			# Log-transform data if specified.
			log_data = [[] for _ in clean_data]
			log_error = False
			if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
				for i in range(len(clean_data[0])):
					try:
						log_data[0].append(math.log10(clean_data[0][i]))
					except:
						log_error = True
						break
				if not log_error:
					clean_data[0] = log_data[0]
					self.data_labels[0] = "Log10 of " + self.x_var.get()
			if self.ylog_ck["state"] != "disabled" and self.ylog_var.get() == "1" and len(clean_data) > 1:
				log_error = False
				for i in range(len(clean_data[1])):
					try:
						log_data[1].append(math.log10(clean_data[1][i]))
					except:
						log_error = True
						break
				if not log_error:
					clean_data[1] = log_data[1]
					self.data_labels[1] = "Log10 of " + self.y_var.get()
			log_data = None
			#
			self.dataset = clean_data
			self.data_btn["state"] = "normal"
			# Summarize and sort the data as needed for each type of plot.
			if plot_type == "Category counts":
				# Count of values for each X, ordered by X
				counter = collections.Counter(self.dataset[0])
				x_vals = list(counter.keys())
				x_vals.sort()
				x_counts = [counter[k] for k in x_vals]
				self.plot_data = [x_vals, x_counts]
				#self.plot_data_labels = [self.x_var.get(), "Count"]
				self.plot_data_labels = [self.data_labels[0], "Count"]
			elif plot_type == "Box plot":
				# A list of Y values for each X value
				x_vals = list(set(self.dataset[0]))
				ds = list(zip(self.dataset[0], self.dataset[1]))
				plot_data = []
				for x in x_vals:
					plot_data.append([d[1] for d in ds if d[0] == x])
				self.plot_data = plot_data
				self.plot_data_labels = x_vals
			elif plot_type == "Empirical CDF":
				# Y is the fraction of data points below each X value
				x_counts = np.unique(self.dataset[0], return_counts=True)
				y_vals = list(np.cumsum(x_counts[1]/np.sum(x_counts[1])))
				self.plot_data = [list(x_counts[0]), y_vals]
				self.plot_data_labels = [self.data_labels[0], "Cumulative frequency"]
			elif plot_type == "Y range plot":
				# Min and max Y for each X
				x_vals = list(set(self.dataset[0]))
				x_vals.sort()
				y_vals = [[None, None]] * len(x_vals)
				plotdata = dict(zip(x_vals, y_vals))
				for i in range(len(self.dataset[0])):
					x = self.dataset[0][i]
					y = self.dataset[1][i]
					y_vals = plotdata[x]
					if y_vals[0] is None or y < y_vals[0]:
						plotdata[x][0] = y
					if y_vals[1] is None or y > y_vals[1]:
						plotdata[x][1] = y
				y1 = [plotdata[x][0] for x in x_vals]
				y2 = [plotdata[x][1] for x in x_vals]
				self.plot_data = [x_vals, y1, y2]
				self.plot_data_labels = [self.x_var.get(), self.y_var.get() + " min", self.y_var.get() + " max"]
			elif plot_type == "Line plot":
				# Sort by X
				ds = list(zip(self.dataset[0], self.dataset[1]))
				ds.sort()
				ds2 = list(zip(*ds))
				self.plot_data = [list(ds2[0]), list(ds2[1])]
				self.plot_data_labels = self.data_labels
			elif plot_type in ("Histogram", "Scatter plot", "Y range plot"):
				# No special preparation
				self.plot_data = self.dataset
				self.plot_data_labels = self.data_labels
			self.plot_data_btn["state"] = "normal"

	def redraw(self):
		#self.plotfig.clear()
		#self.plot_axes = self.plotfig.add_subplot(111)
		#self.plotfig_canvas.draw()
		plot_type = self.type_var.get()
		if self.plot_data is not None and len(self.plot_data[0]) > 0:
			if plot_type == "Category counts":
				self.plot_axes.bar(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Histogram":
				self.plot_axes.hist(self.plot_data[0])
				self.plot_axes.set_xlabel(self.x_var.get())
				self.plot_axes.set_ylabel("Counts")
			elif plot_type == "Scatter plot":
				self.plot_axes.scatter(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Line plot":
				self.plot_axes.plot(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Empirical CDF":
				self.plot_axes.stackplot(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Y range plot":
				self.plot_axes.fill_between(self.plot_data[0], self.plot_data[1], self.plot_data[2])
				self.plot_axes.set_xlabel(self.x_var.get())
				self.plot_axes.set_ylabel(self.y_var.get())
			elif plot_type == "Box plot":
				self.plot_axes.boxplot(self.plot_data, labels=self.plot_data_labels)
				self.plot_axes.set_xlabel(self.x_var.get())
				self.plot_axes.set_ylabel(self.data_labels[1])
			self.plotfig_canvas.draw()
			self.plot_nav.update()

	def do_close(self, *args):
		self.parent.remove_plot(self)
		self.dlg.destroy()
	def show(self):
		self.dlg.update_idle_tasks()
		self.dlg.minsize(width=500, height=500)
		self.dlg.wait_window(self.dlg)



class MsgDialog(object):
	#def __init__(self, title, message, width=400, height=400):
	def __init__(self, title, message):
		self.dlg = tk.Toplevel()
		self.dlg.title(title)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		msg_lbl = ttk.Label(prompt_frame, text=message)
		msg_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		self.content_frame = tk.Frame(self.dlg)
		self.content_frame.grid(row=1, column=0, sticky=tk.NSEW)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		ok_btn = ttk.Button(btn_frame, text="Close", command=self.do_select)
		ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_select)
	def do_select(self, *args):
		self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)



class SelDataSrcDialog(object):
	#def __init__(self, title, message, width=400, height=400):
	def __init__(self):
		self.canceled = False
		self.dlg = tk.Toplevel()
		self.dlg.title("Select Mapping Data")
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.dlg.columnconfigure(0, weight=1)
		self.rv = (None, None, None, None, None, None, None, None, None, None)
		# Prompt
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,6), pady=(6,3))
		msg_lbl = ttk.Label(prompt_frame, width=30, wraplength=100, anchor=tk.W, justify=tk.LEFT, text="Select the type of data source to use.  You will then be prompted for details about the selected data source.")
		msg_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		# Select buttons
		sel_frame = tk.Frame(self.dlg)
		sel_frame.grid(row=1, column=0, sticky=tk.NSEW, pady=(6,9))
		csv_btn = ttk.Button(sel_frame, text=" CSV file  ", command=self.sel_csv)
		csv_btn.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		ss_btn = ttk.Button(sel_frame,  text="Spreadsheet", command=self.sel_spreadsheet)
		ss_btn.grid(row=0, column=1, sticky=tk.EW, padx=(3,3), pady=(3,3))
		db_btn = ttk.Button(sel_frame,  text=" Database  ", command=self.sel_database)
		db_btn.grid(row=0, column=2, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Help and Cancel buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=0, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.osdn.io/", new=2, autoraise=True)
	def sel_csv(self):
		dfd = DataFileDialog()
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_spreadsheet(self):
		dfd = ImportSpreadsheetDialog()
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_database(self):
		dbd = DbConnectDialog()
		self.rv = dbd.get_data()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def select(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		return self.rv



class EncodedFile(object):
	# A class providing an open method for an encoded file, allowing reading
	# and writing using unicode, without explicit decoding or encoding.
	def __repr__(self):
		return u"EncodedFile(%r, %r)" % (self.filename, self.encoding)
	def __init__(self, filename, file_encoding):
		self.filename = filename
		self.encoding = file_encoding
		self.bom_length = 0
		def detect_by_bom(path, default_enc):
			with io.open(path, 'rb') as f:
				raw = f.read(4)
			for enc, boms, bom_len in (
							('utf-8-sig', (codecs.BOM_UTF8,), 3),
							('utf_16', (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE), 2),
							('utf_32', (codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE), 4)):
				if any(raw.startswith(bom) for bom in boms):
					return enc, bom_len
			return default_enc, 0
		if os.path.exists(filename):
			self.encoding, self.bom_length = detect_by_bom(filename, file_encoding)
		self.fo = None
	def open(self, mode='r'):
		self.fo = io.open(file=self.filename, mode=mode, encoding="UTF8", newline=None)
		return self.fo
	def close(self):
		if self.fo is not None:
			self.fo.close()


class LineDelimiter(object):
	def __init__(self, delim, quote, escchar):
		self.delimiter = delim
		self.joinchar = delim if delim else u""
		self.quotechar = quote
		if quote:
			if escchar:
				self.quotedquote = escchar+quote
			else:
				self.quotedquote = quote+quote
		else:
			self.quotedquote = None
	def delimited(self, datarow, add_newline=True):
		global conf
		if self.quotechar:
			d_row = []
			for e in datarow:
				if isinstance(e, str):
					if (self.quotechar in e) or (self.delimiter is not None and self.delimiter in e) or (u'\n' in e) or (u'\r' in e):
						d_row.append(u"%s%s%s" % (self.quotechar, e.replace(self.quotechar, self.quotedquote), self.quotechar))
					else:
						d_row.append(e)
				else:
					if e is None:
						d_row.append('')
					else:
						d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		else:
			d_row = []
			for e in datarow:
				if e is None:
					d_row.append('')
				else:
					d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		if add_newline:
			text = text + u"\n"
		return text


def write_delimited_file(outfile, filefmt, column_headers, rowsource, file_encoding='utf8', append=False):
	delim = None
	quote = None
	escchar = None
	if filefmt.lower() == 'csv':
		delim = ","
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('tab', 'tsv'):
		delim = "\t"
		quote = None
		escchar = None
	elif filefmt.lower() in ('tabq', 'tsvq'):
		delim = "\t"
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('unitsep', 'us'):
		delim = chr(31)
		quote = None
		escchar = None
	elif filefmt.lower() == 'plain':
		delim = " "
		quote = ''
		escchar = None
	elif filefmt.lower() == 'tex':
		delim = "&"
		quote = ''
		escchar = None
	line_delimiter = LineDelimiter(delim, quote, escchar)
	fmode = "w" if not append else "a"
	ofile = EncodedFile(outfile, file_encoding).open(mode=fmode)
	fdesc = outfile
	if not (filefmt.lower() == 'plain' or append):
		datarow = line_delimiter.delimited(column_headers)
		ofile.write(datarow)
	for rec in rowsource:
		datarow = line_delimiter.delimited(rec)
		ofile.write(datarow)
	ofile.close()



class OdsFile(object):
	def __repr__(self):
		return u"OdsFile()"
	def __init__(self):
		self.filename = None
		self.wbk = None
		self.cell_style_names = []
	def open(self, filename):
		self.filename = filename
		if os.path.isfile(filename):
			self.wbk = odf.opendocument.load(filename)
			# Get a list of all cell style names used, so as not to re-define them.
			# Adapted from http://www.pbertrand.eu/reading-an-odf-document-with-odfpy/
			for sty in self.wbk.automaticstyles.childNodes:
				try:
					fam = sty.getAttribute("family")
					if fam == "table-cell":
						name = sty.getAttribute("name")
						if not name in self.cell_style_names:
							self.cell_style_names.append(name)
				except:
					pass
		else:
			self.wbk = odf.opendocument.OpenDocumentSpreadsheet()
	def define_body_style(self):
		st_name = "body"
		if not st_name in self.cell_style_names:
			body_style = odf.style.Style(name=st_name, family="table-cell")
			body_style.addElement(odf.style.TableCellProperties(attributes={"verticalalign":"top"}))
			self.wbk.styles.addElement(body_style)
			self.cell_style_names.append(st_name)
	def define_header_style(self):
		st_name = "header"
		if not st_name in self.cell_style_names:
			header_style = odf.style.Style(name=st_name, family="table-cell")
			header_style.addElement(odf.style.TableCellProperties(attributes={"borderbottom":"1pt solid #000000",
				"verticalalign":"bottom"}))
			self.wbk.styles.addElement(header_style)
			self.cell_style_names.append(st_name)
	def define_iso_datetime_style(self):
		st_name = "iso_datetime"
		if not st_name in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-datetime")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			# odfpy collapses text elements that have only spaces, so trying to insert just a space between the date
			# and time actually results in no space between them.  Other Unicode invisible characters
			# are also trimmed.  The delimiter "T" is used instead, and conforms to ISO-8601 specifications.
			dt_style.addElement(odf.number.Text(text=u"T"))
			dt_style.addElement(odf.number.Hours(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Minutes(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Seconds(style="long", decimalplaces="3"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-datetime", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def define_iso_date_style(self):
		st_name = "iso_date"
		if st_name not in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-date")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-date", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def sheetnames(self):
		# Returns a list of the worksheet names in the specified ODS spreadsheet.
		return [sheet.getAttribute("name") for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table)]
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			for i, sheet in enumerate(self.wbk.spreadsheet.getElementsByType(odf.table.Table)):
				if i+1 == sheet_no:
					return sheet
			else:
				sheet_no = None
		if sheet_no is None:
			for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table):
				if sheet.getAttribute("name").lower() == sheetname.lower():
					return sheet
		return None
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		if not sheet:
			warning("There is no sheet named %s" % sheetname, kwargs={})
			raise
		def row_data(sheetrow):
			# Adapted from http://www.marco83.com/work/wp-content/uploads/2011/11/odf-to-array.py
			cells = sheetrow.getElementsByType(odf.table.TableCell)
			rowdata = []
			for cell in cells:
				p_content = []
				repeat = cell.getAttribute("numbercolumnsrepeated")
				if not repeat:
					repeat = 1
					spanned = int(cell.getAttribute("numbercolumnsspanned") or 0)
					if spanned > 1:
						repeat = spanned
				ps = cell.getElementsByType(odf.text.P)
				if len(ps) == 0:
					for rr in range(int(repeat)):
						p_content.append(None)
				else:
					for p in ps:
						pval = type(u"")(p)
						if len(pval) == 0:
							for rr in range(int(repeat)):
								p_content.append(None)
						else:
							for rr in range(int(repeat)):
								p_content.append(pval)
				if len(p_content) == 0:
					for rr in range(int(repeat)):
						rowdata.append(None)
				elif p_content[0] != u'#':
					rowdata.extend(p_content)
			return rowdata
		rows = sheet.getElementsByType(odf.table.TableRow)
		if junk_header_rows > 0:
			rows = rows[junk_header_rows: ]
		return [row_data(r) for r in rows]
	def new_sheet(self, sheetname):
		# Returns a sheet (a named Table) that has not yet been added to the workbook
		return odf.table.Table(name=sheetname)
	def add_row_to_sheet(self, datarow, odf_table, header=False):
		if header:
			self.define_header_style()
			style_name = "header"
		else:
			self.define_body_style()
			style_name = "body"
		tr = odf.table.TableRow()
		odf_table.addElement(tr)
		for item in datarow:
			if isinstance(item, bool):
				# Booleans must be evaluated before numbers.
				# Neither of the first two commented-out lines actually work (a bug in odfpy?).
				# Booleans *can* be written as either integers or strings; integers are chosen below.
				#tc = odf.table.TableCell(booleanvalue='true' if item else 'false')
				#tc = odf.table.TableCell(valuetype="boolean", value='true' if item else 'false')
				tc = odf.table.TableCell(valuetype="boolean", value=1 if item else 0, stylename=style_name)
				#tc = odf.table.TableCell(valuetype="string", stringvalue='True' if item else 'False')
			elif isinstance(item, float) or isinstance(item, int):
				tc = odf.table.TableCell(valuetype="float", value=item, stylename=style_name)
			elif isinstance(item, datetime.datetime):
				self.define_iso_datetime_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3], stylename="iso_datetime")
			elif isinstance(item, datetime.date):
				self.define_iso_date_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%d"), stylename="iso_date")
			elif isinstance(item, datetime.time):
				self.define_iso_datetime_style()
				timeval = datetime.datetime(1899, 12, 30, item.hour, item.minute, item.second, item.microsecond, item.tzinfo)
				tc = odf.table.TableCell(timevalue=timeval.strftime("PT%HH%MM%S.%fS"), stylename="iso_datetime")
				tc.addElement(odf.text.P(text=timeval.strftime("%H:%M:%S.%f")))
			elif isinstance(item, str):
				item = item.replace(u'\n', u' ').replace(u'\r', u' ')
				tc = odf.table.TableCell(valuetype="string", stringvalue=item, stylename=style_name)
			else:
				tc = odf.table.TableCell(value=item, stylename=style_name)
			if item is not None:
				tc.addElement(odf.text.P(text=item))
			tr.addElement(tc)
	def add_sheet(self, odf_table):
		self.wbk.spreadsheet.addElement(odf_table)
	def save_close(self):
		ofile = io.open(self.filename, "wb")
		self.wbk.write(ofile)
		ofile.close()
		self.filename = None
		self.wbk = None
	def close(self):
		self.filename = None
		self.wbk = None


def ods_data(filename, sheetname, junk_header_rows=0):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	wbk = OdsFile()
	try:
		wbk.open(filename)
	except:
		warning("%s is not a valid OpenDocument spreadsheet." % filename, kwargs={})
		raise
	try:
		alldata = wbk.sheet_data(sheetname, junk_header_rows)
	except:
		warning("%s is not a worksheet in %s." % (sheetname, filename), kwargs={})
		raise
	colhdrs = alldata[0]
	if any([x is None or len(x.strip())==0 for x in colhdrs]):
		if conf.del_empty_cols:
			blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
			while len(blanks) > 0:
				b = blanks.pop()
				for r in range(len(alldata)):
					del(alldata[r][b])
			colhdrs = alldata[0]
		else:
			if conf.create_col_hdrs:
				for i in range(len(colhdrs)):
					if colhdrs[i] is None or len(colhdrs[i]) == 0:
						colhdrs[i] = "Col%s" % str(i+1)
			else:
				warning("The input file %s, sheet %s has missing column headers." % (filename, sheetname), kwargs={})
				raise
	#if conf.clean_col_hdrs:
	#	colhdrs = clean_words(colhdrs)
	#if conf.trim_col_hdrs != 'none':
	#	colhdrs = trim_words(colhdrs, conf.trim_col_hdrs)
	#if conf.fold_col_hdrs != 'no':
	#	colhdrs = fold_words(colhdrs, conf.fold_col_hdrs)
	#if conf.dedup_col_hdrs:
	#	colhdrs = dedup_words(colhdrs)
	return colhdrs, alldata[1:]


def export_ods(outfile, hdrs, rows, append=False, querytext=None, sheetname=None, desc=None):
	# If not given, determine the worksheet name to use.  The pattern is "Sheetx", where x is
	# the first integer for which there is not already a sheet name.
	if append and os.path.isfile(outfile):
		wbk = OdsFile()
		wbk.open(outfile)
		sheet_names = wbk.sheetnames()
		name = sheetname or u"Sheet"
		sheet_name = name
		sheet_no = 1
		while True:
			if sheet_name not in sheet_names:
				break
			sheet_no += 1
			sheet_name = u"%s%d" % (name, sheet_no)
		wbk.close()
	else:
		sheet_name = sheetname or u"Sheet1"
		if os.path.isfile(outfile):
			os.unlink(outfile)
	wbk = OdsFile()
	wbk.open(outfile)
	# Add the data to a new sheet.
	tbl = wbk.new_sheet(sheet_name)
	wbk.add_row_to_sheet(hdrs, tbl, header=True)
	for row in rows:
		wbk.add_row_to_sheet(row, tbl)
	# Add sheet to workbook
	wbk.add_sheet(tbl)
	# Save and close the workbook.
	wbk.save_close()


class XlsFile(object):
	def __repr__(self):
		return u"XlsFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.datemode = 0
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		self.wbk = xlrd.open_workbook(filename, encoding_override=self.encoding)
		self.datemode = self.wbk.datemode
	def sheetnames(self):
		return self.wbk.sheet_names()
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is None:
			sheet = self.wbk.sheet_by_name(sheetname)
		else:
			# User-specified sheet numbers should be 1-based; xlrd sheet indexes are 0-based
			sheet = self.wbk.sheet_by_index(max(0, sheet_no-1))
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.ncols and sheet.nrows, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		def row_data(sheetrow, columns=None):
			cells = sheet.row_slice(sheetrow)
			if columns:
				d = [cells[c] for c in range(columns)]
			else:
				d = [cell for cell in cells]
			datarow = []
			for c in d:
				if c.ctype == 0:
					# empty
					datarow.append(None)
				elif c.ctype == 1:
					datarow.append(c.value)
				elif c.ctype == 2:
					# float, but maybe should be int
					if c.value - int(c.value) == 0:
						datarow.append(int(c.value))
					else:
						datarow.append(c.value)
				elif c.ctype == 3:
					# date
					dt = xlrd.xldate_as_tuple(c.value, self.datemode)
					# Convert to time or datetime
					if not any(dt[:3]):
						# No date values
						datarow.append(datetime.time(*dt[3:]))
					else:
						datarow.append(datetime.datetime(*dt))
				elif c.ctype == 4:
					# Boolean
					datarow.append(bool(c.value))
				elif c.ctype == 5:
					# Error code
					datarow.append(xlrd.error_text_from_code(c.value))
				elif c.ctype == 6:
					# blank
					datarow.append(None)
				else:
					datarow.append(c.value)
			return datarow
		hdr_row = row_data(junk_header_rows)
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		sheet_data = []
		for r in range(junk_header_rows, sheet.nrows - junk_header_rows):
			datarow = row_data(r, ncols)
			if datarow.count(None) == len(datarow):
				break
			sheet_data.append(datarow)
		return sheet_data


class XlsxFile(object):
	def __repr__(self):
		return u"XlsxFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.read_only = False
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		if os.path.isfile(filename):
			if read_only:
				self.wbk = openpyxl.load_workbook(filename, read_only=True)
			else:
				self.wbk = openpyxl.load_workbook(filename)
	def close(self):
		if self.wbk is not None:
			self.wbk.close()
			self.wbk = None
			self.filename = None
			self.encoding = None
	def sheetnames(self):
		return self.wbk.sheetnames
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			# User-specified sheet numbers should be 1-based
			sheet = self.wbk[self.wbk.sheetnames[sheet_no - 1]]
		else:
			sheet = self.wbk[sheetname]
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.max_column and sheet.max_row, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		# Get the header row, skipping junk rows
		rowsrc = sheet.iter_rows(max_row = junk_header_rows + 1, values_only = True)
		for hdr_row in rowsrc:
			pass
		# Get the number of columns
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		# Get all the data rows
		sheet_data = []
		rowsrc = sheet.iter_rows(min_row = junk_header_rows + 1, values_only = True)
		for r in rowsrc:
			if not any(r):
				break
			sheet_data.append(list(r))
		for r in range(len(sheet_data)):
			rd = sheet_data[r]
			for c in range(len(rd)):
				if isinstance(rd[c], str):
					if rd[c] == '=FALSE()':
						rd[c] = False
					elif rd[c] == '=TRUE()':
						rd[c] = True
		return sheet_data


def xls_data(filename, sheetname, junk_header_rows, encoding=None):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	root, ext = os.path.splitext(filename)
	ext = ext.lower()
	if ext == ".xls":
		wbk = XlsFile()
	else:
		wbk = XlsxFile()
	try:
		wbk.open(filename, encoding, read_only=True)
	except:
		warning("%s is not a valid Excel spreadsheet." % filename, kwargs={})
		raise
	alldata = wbk.sheet_data(sheetname, junk_header_rows)
	if len(alldata) == 0:
		raise ErrInfo(type="cmd", other_msg="There are no data on worksheet %s of file %s." % (sheetname, filename))
	if ext == 'xlsx':
		wbk.close()
	if len(alldata) == 1:
		return alldata[0], []
	colhdrs = alldata[0]
	# Delete columns with missing headers
	if any([x is None or (isinstance(x, str) and len(x.strip())==0) for x in colhdrs]):
		blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
		while len(blanks) > 0:
			b = blanks.pop()
			for r in range(len(alldata)):
				del(alldata[r][b])
		colhdrs = alldata[0]
	#if conf.clean_col_hdrs:
	#	colhdrs = clean_words(colhdrs)
	#if conf.trim_col_hdrs != 'none':
	#	colhdrs = trim_words(colhdrs, conf.trim_col_hdrs)
	#if conf.fold_col_hdrs != 'no':
	#	colhdrs = fold_words(colhdrs, conf.fold_col_hdrs)
	#if conf.dedup_col_hdrs:
	#	colhdrs = dedup_words(colhdrs)
	return colhdrs, alldata[1:]



def file_data(filename):
	# Get headers and rows from the specified CSV file
	csvreader = CsvFile(filename)
	headers = csvreader.next()
	rows = []
	for line in csvreader:
		rows.append(line)
	return headers, rows


def read_all_config(datafile=None):
	global config_files
	config_files = []
	if os.name == 'posix':
		sys_config_file = os.path.join("/etc", config_file_name)
	else:
		sys_config_file = os.path.join(os.path.expandvars(r'%APPDIR%'), config_file_name)
	if os.path.isfile(sys_config_file):
		config_files.append(sys_config_file)
	program_dir_config = os.path.join(os.path.abspath(sys.argv[0]), config_file_name)
	if os.path.isfile(program_dir_config) and not program_dir_config in config_files:
		config_files.append(program_dir_config)
	user_config_file = os.path.join(os.path.expanduser(r'~/.config'), config_file_name)
	if os.path.isfile(user_config_file) and not user_config_file in config_files:
		config_files.append(user_config_file)
	if datafile is not None:
		data_config_file = os.path.join(os.path.abspath(datafile), config_file_name)
		if os.path.isfile(data_config_file) and not data_config_file in config_files:
			config_files.append(data_config_file)
	startdir_config_file = os.path.join(os.path.abspath(os.path.curdir), config_file_name)
	if os.path.isfile(startdir_config_file) and not startdir_config_file in config_files:
		config_files.append(startdir_config_file)
	files_read = []
	for config_file in config_files:
		files_read.append(config_file)
		read_config(config_file)


def read_config(configfile):
	_BASEMAP_SECTION = "basemap_tile_servers"
	_APIKEYS_SECTION = "api_keys"
	_SYMBOL_SECTION = "symbols"
	_DEFAULTS_SECTION = "defaults"
	_MISC_SECTION = "misc"
	cp = ConfigParser()
	cp.read(configfile)
	# Tile servers
	if cp.has_section(_BASEMAP_SECTION):
		basemap_sources = cp.items(_BASEMAP_SECTION)
		for name, url in basemap_sources:
			if url is None:
				if name in bm_servers and len(bm_servers) > 1:
					del(bm_servers[name])
			else:
				bm_servers[name.capitalize()] = url
	# API keys
	if cp.has_section(_APIKEYS_SECTION):
		apikeys = cp.items(_APIKEYS_SECTION)
		for name, apikey in apikeys:
			if apikey is None:
				if name in api_keys and len(api_keys) > 1:
					del(api_keys[name])
			else:
				api_keys[name.capitalize()] = apikey
	# Symbols
	if cp.has_section(_SYMBOL_SECTION):
		symbols = cp.items(_SYMBOL_SECTION)
		for name, filename in symbols:
			import_symbol(name, filename)
	# Defaults
	if cp.has_option(_DEFAULTS_SECTION, "multiselect"):
		global multiselect
		err = False
		try:
			multi = cp.getboolean(_DEFAULTS_SECTION, "multiselect")
		except:
			err = True
			warning("Invalid argument to the 'multiselect' configuration option", kwargs={})
		if not err:
			multiselect = "1" if multi else "0"
	if cp.has_option(_DEFAULTS_SECTION, "basemap"):
		global initial_basemap
		bm = cp.get(_DEFAULTS_SECTION, "basemap")
		if bm is None or bm not in bm_servers:
			warning("Invalid argument to the 'basemap' configuration option", kwargs={})
		else:
			initial_basemap = bm
	if cp.has_option(_DEFAULTS_SECTION, "location_marker"):
		global location_marker
		loc_mkr = cp.get(_DEFAULTS_SECTION, "location_marker")
		if loc_mkr is not None:
			location_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "location_color"):
		global location_color
		loc_color = cp.get(_DEFAULTS_SECTION, "location_color")
		if loc_color is not None:
			if loc_color not in color_names:
				warning("Invalid argument to the 'location_color' configuration option", kwargs={})
			else:
				location_color = loc_color
	if cp.has_option(_DEFAULTS_SECTION, "use_data_marker"):
		global use_data_marker
		loc_mkr = cp.getboolean(_DEFAULTS_SECTION, "use_data_marker")
		if loc_mkr is not None:
			use_data_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "use_data_color"):
		global use_data_color
		loc_clr = cp.getboolean(_DEFAULTS_SECTION, "use_data_color")
		if loc_clr is not None:
			use_data_color = loc_clr
	if cp.has_option(_DEFAULTS_SECTION, "select_symbol"):
		global select_symbol
		default_symbol = cp.get(_DEFAULTS_SECTION, "select_symbol")
		if default_symbol is not None:
			if default_symbol not in icon_xbm:
				warning("Unrecognized symbol name for the 'select_symbol' configuration option", kwargs={})
			else:
				select_symbol = default_symbol
	if cp.has_option(_DEFAULTS_SECTION, "select_color"):
		global select_color
		sel_color = cp.get(_DEFAULTS_SECTION, "select_color")
		if sel_color is not None:
			if sel_color not in color_names:
				warning("Invalid argument to the 'multiselect' configuration option", kwargs={})
			else:
				select_color = sel_color
	if cp.has_option(_DEFAULTS_SECTION, "label_color"):
		global label_color
		lbl_color = cp.get(_DEFAULTS_SECTION, "label_color")
		if lbl_color is not None:
			if lbl_color not in color_names:
				warning("Invalid argument to the 'label_color' configuration option", kwargs={})
			else:
				label_color = lbl_color
	if cp.has_option(_DEFAULTS_SECTION, "label_font"):
		global label_font
		lbl_font = cp.get(_DEFAULTS_SECTION, "label_font")
		if lbl_font is not None:
			if lbl_font not in list(tk.font.families()):
				warning("Invalid argument to the 'label_font' configuration option", kwargs={})
			else:
				label_font = lbl_font
	if cp.has_option(_DEFAULTS_SECTION, "label_size"):
		global label_size
		err = False
		try:
			lbl_size = cp.getint(_DEFAULTS_SECTION, "label_size")
		except:
			err = True
			warning("Invalid argument to the 'label_size' configuration option", kwargs={})
		if not err:
			if lbl_size is not None and lbl_size > 6:
				label_size = lbl_size
	if cp.has_option(_DEFAULTS_SECTION, "label_bold"):
		global label_bold
		err = False
		try:
			lbl_bold = cp.getboolean(_DEFAULTS_SECTION, "label_bold")
		except:
			err = True
			warning("Invalid argument to the 'label_bold' configuration option", kwargs={})
		if not err:
			if lbl_bold is not None:
				label_bold = lbl_bold
	if cp.has_option(_DEFAULTS_SECTION, "label_position"):
		global label_position
		lbl_position = cp.get(_DEFAULTS_SECTION, "label_position")
		if lbl_position is not None:
			lbl_position = lbl_position.lower()
			if lbl_position not in ("above", "below"):
				warning("Invalid argument to the 'label_position' configuration option", kwargs={})
			else:
				label_position = lbl_position
	if cp.has_option(_MISC_SECTION, "temp_dbfile"):
		global temp_dbfile
		err = False
		try:
			dbfile = cp.getboolean(_MISC_SECTION, "temp_dbfile")
		except:
			err = True
			warning("Invalid argument to the 'temp_dbfile' configuration option", kwargs={})
		if not err:
			temp_dbfile = dbfile
	if cp.has_option(_MISC_SECTION, "editor"):
		global editor
		err = False
		try:
			ed = cp.get(_MISC_SECTION, "editor")
		except:
			err = True
			warning("Invalid argument to the 'editor' configuration option", kwargs={})
		if not err:
			editor = ed



def import_symbol(symbol_name, filename):
	with open(filename, mode='r') as f:
		symbol_def = f.read()
	icon_xbm[symbol_name] = symbol_def




def clparser():
	desc_msg = "Display an interactive map with points read from a CSV file. Version %s, %s" % (version, vdate)
	parser = argparse.ArgumentParser(description=desc_msg)
	parser.add_argument('-f', '--file', default=None,
			help="The name of a CSV or spreadsheet file containing latitude and longitude coordinates")
	parser.add_argument('-m', '--message',
			dest='message', default='Map display.',
			help='A message to display above the map')
	parser.add_argument('-t', '--sheet', default=None,
			help="The name of a worksheet when the data source is a spreadsheet")
	parser.add_argument('-i', '--identifier', default='location_id', dest='id',
			help="The name of the column in the data file containing location identifiers or labels (default: location_id)")
	parser.add_argument('-x', '--lon', default='x_coord', dest='lon',
			help="The name of the column in the data file containing longitude values (default: x_coord)")
	parser.add_argument('-y', '--lat', default='y_coord', dest='lat',
			help="The name of the column in the data file containg latitude values (default: y_coord)")
	parser.add_argument('-s', '--symbol', default=None, dest='symbol',
			help="The name of the column in the data file containing symbol names")
	parser.add_argument('-c', '--color', default=None, dest='color',
			help="The name of the column in the data file containing color names")
	parser.add_argument('-p', '--projection', default=4326,
			help="The coordinate reference system (CRS) if the data are projected (default: 4326, i.e., no projection)")
	parser.add_argument('-g', '--image', dest='imagefile', default=None,
			help="The name of an image file to which the map will be exported--no UI will be created.")
	parser.add_argument('-w', '--imagewait', default=12,
			help="The time in seconds to wait before exporting the map to an image file.")
	return parser



def main():
	args = clparser().parse_args()
	if args.file is None or args.lat is None or args.lon is None:
		fn = lat_col = lon_col = id_col = sym_col = col_col = crs = sheet = msg = headers = rows = imagefile = None
		imagewait = 12
	else:
		fn = args.file
		if not os.path.exists(fn):
			fatal_error("File %s does not exist" % fn)
		sheet = args.sheet
		lat_col = args.lat
		lon_col = args.lon
		id_col = args.id
		sym_col = args.symbol
		col_col = args.color
		crs = args.projection
		msg = args.message
		imagefile = args.imagefile
		imagewait = args.imagewait
	read_all_config(fn)
	app = MapUI(fn, msg, lat_col, lon_col, crs, sheet, id_col, sym_col, col_col, map_export_file=imagefile,
			export_time_sec=imagewait)
	app.win.mainloop()


main()


