import abc
from typing import Iterator, Optional, Generator, Union
import logging
import csv
import numpy as np
import re
from enum import Enum
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from .bounding_box import BoundingBox
from .point import Point
from .ligatures import LIGATURES_TO_TEXT
from .constants import (
    PAGE_DELIMITER,
    AREA_DELIMITER,
    WORD_DELIMITER,
    LINE_DELIMITER,
    TABLE_DELIMITER,
    PARAGRAPH_DELIMITER,
    CELL_DELIMITER,
    COLUMN_DELIMITER,
    CHARACTER_WIDTH_MAP,
    THIN_CHARS,
    NORMALIZED_BBOX,
)
from .version import VERSION

"""
This module contains the TextBlock class and its subclasses.
The TextBlock class is the base class for all text blocks: Document, Page, Paragraph, Line, Word, Character.
"""


class TextBlock(metaclass=abc.ABCMeta):
    """
    Abstract class for all text blocks: Document, Page, Paragraph, Line, Word, Character.
    Every text block has a bounding box (1), a delimiter (2), a list of children(3) and a parent(4).
    The Document block is the root of the tree, while the Character block is the leaf of the tree.

    (1,4) The Document doesn't have a bounding box nor a parent.
    (2,3) The Character doesn't have a delimiter nor children.
    """

    child_type: type = None
    delimiter: str = ""

    def __init__(
        self,
        bounding_box: BoundingBox = None,
        children: list["TextBlock"] = None,
    ):
        self.bounding_box: BoundingBox = bounding_box
        self.length: Optional[int] = None
        self.parent: Optional["TextBlock"] = None
        self.relative_order: int = None
        self.start_index: int = None
        self.end_index: int = None
        self.set_children(children)
        self.set_bounding_box()

    def get_intersecting_text_blocks(
        self, _class: type, bbox: BoundingBox
    ) -> Generator:
        if isinstance(self, _class) and self.bounding_box.intersect(bbox):
            yield self

        for child in self.children:
            yield from child.get_intersecting_text_blocks(_class, bbox)

    def set_relative_order(self, index: int):
        """
        Set the relative order of the block.
        The relative order of a block is the order of the block
        in the list of its parent's children"""
        self.relative_order = index

    def get_relative_order(self) -> int:
        """
        Return the relative order of the block.
        The relative order of a block is the order of the block
        in the list of its parent's children.
        """
        if self.relative_order is None:
            raise ValueError("The relative order of the block is not set")
        return self.relative_order

    def set_bounding_box(self):
        """
        Set the bounding box by composing the bounding boxes of its children.
        The bounding box of a block is the "minimal" bounding box
        surrounding the text of the block.
        For more information see the documentation of the compose_bounding_boxes
        method at the BoundingBox class.
        """
        if self.bounding_box is not None:
            return
        bounding_boxes = [child.get_bounding_box() for child in self.children]
        for bounding_box in bounding_boxes:
            if not isinstance(bounding_box, BoundingBox):
                raise TypeError(
                    (
                        "When calling set_bounding_box for a box, "
                        "all of it's children must have a bounding box"
                    )
                )
        self.bounding_box = BoundingBox.compose_bounding_boxes(bounding_boxes)

    def get_bounding_box(self) -> BoundingBox:
        """
        Return the bounding box of a text block.
        The bounding box of the box is the "minimal" bounding box
        surrounding the text of the block."""
        if self.bounding_box is None:
            self.set_bounding_box()
        return self.bounding_box

    def set_children(self, children: Optional[list["TextBlock"]]):
        """Set the children of the text block."""
        self.children: list[self.child_type] = []
        if children is None:
            return
        for child in children:
            self.set_child(child)

    def get_children(self) -> list:
        """
        Return the children of the text block."""
        return self.children

    def get_parent(self) -> "TextBlock":
        """
        Returns the parent of the block.
        """
        return self.parent

    def set_parent(self, parent: Optional["TextBlock"]):
        """Set the parent of the text block."""
        self.parent = parent

    @staticmethod
    def sort(
        text_blocks: list["TextBlock"],
        return_indexes: bool = False,
        height_scale: float = 1,
    ) -> list["TextBlock"]:
        """
        Sort the list of text blocks by their relative order.
        """
        sorted_indexes = TextBlock._sort(
            text_blocks, height_scale=height_scale
        )
        if return_indexes:
            return sorted_indexes
        return [text_blocks[index] for index in sorted_indexes]

    @staticmethod
    def _sort(
        text_blocks: list["TextBlock"], height_scale: float
    ) -> list[int]:
        indexes = list(range(len(text_blocks)))
        if len(text_blocks) <= 1:
            return indexes
        centers = [
            text_block.bounding_box.get_center() for text_block in text_blocks
        ]
        sorted_indexes = sorted(
            indexes, key=lambda index: centers[index].x, reverse=False
        )
        sorted_indexes = sorted(
            sorted_indexes, key=lambda index: centers[index].y, reverse=True
        )
        indexes_groups = []
        indexes_group = [sorted_indexes[0]]
        for i in range(len(sorted_indexes) - 1):
            first_bbox = text_blocks[sorted_indexes[i]].bounding_box.scale(
                height_scale=height_scale
            )
            second_bbox = text_blocks[
                sorted_indexes[i + 1]
            ].bounding_box.scale(height_scale=height_scale)
            if first_bbox.vertical_intersect(second_bbox):
                indexes_group.append(sorted_indexes[i + 1])
            else:
                indexes_groups.append(indexes_group)
                indexes_group = [sorted_indexes[i + 1]]
        indexes_groups.append(indexes_group)
        sorted_indexes = []
        for indexes_group in indexes_groups:
            sorted_indexes.extend(
                sorted(indexes_group, key=lambda index: centers[index].x)
            )
        return sorted_indexes

    @staticmethod
    def pre_order_traversal(root: Optional["TextBlock"]) -> Generator:
        """
        Pre-order traversal is a depth-first traversal
        where the root is visited first, then the children.
        """
        if root is None:
            return
        yield root
        for child in root.children:
            yield from TextBlock.pre_order_traversal(child)

    @staticmethod
    def post_order_traversal(root: Optional["TextBlock"]) -> Generator:
        """
        Post-order traversal is a depth-first traversal where
        the root is visited last, after the children.
        """
        if root is None:
            return
        for child in root.children:
            yield from TextBlock.post_order_traversal(child)

        yield root

    def get_ancestor(self, _class: type) -> Optional["TextBlock"]:
        """
        Returns the ancestor of the block based on the type of the ancestor.
        """
        ancestor = self
        while ancestor is not None and not isinstance(ancestor, _class):
            ancestor = ancestor.parent
        return ancestor

    def set_child(self, child: "TextBlock"):
        """
        Adding a child to the list of children of the block.
        The parent and the relative order of the child is set accordingly.
        """
        if not isinstance(child, self.child_type):
            raise TypeError(
                f"Child of type {type(self)} must be of type {self.child_type}"
            )
        self.children.append(child)
        child.set_parent(self)
        child.set_relative_order(len(self.children) - 1)

    def get_text(self):
        """
        Every block has an implicit text representation that can be
        retrieved by calling this method.
        """
        if getattr(self, "text", None) is not None:
            return self.text
        if not self.children:
            return ""
        delimiter = self.children[0].delimiter
        return delimiter.join([child.get_text() for child in self.children])

    def get_length(self) -> int:
        """Return the length of the text block.
        The length of a text block is equal to the length of the text it represents.
        """
        if self.length is None:
            self.set_length()
        return self.length

    def set_length(self):
        """
        Set the length of the text block.
        The length of a text block is equal to the length of the text it represents.
        """
        if self.length is not None:
            return
        if not self.children:
            self.length = 0
            return
        delimiter = self.child_type.delimiter
        lengths = [child.get_length() for child in self.children]
        self.length = sum(lengths) + len(delimiter) * (len(self.children) - 1)

    def get_all(self, _class: type) -> Generator:
        """
        Returns a generator of all the blocks of the given type that are descendants of the block.
        """
        if issubclass(type(self), _class):
            yield self
        for child in self.children:
            yield from child.get_all(_class)

    def __str__(self):
        return self.get_text()

    def __repr__(self):
        return self.get_text()

    def __len__(self):
        """The length of a block is the number of characters it contains, including the delimiters.
        The length is equivalent to len(block.get_text())"""
        return self.length

    def __iter__(self) -> Iterator["Character"]:
        """Iterating over the characters of the block"""
        return iter(self.get_all(Character))

    def __getitem__(self, index: Union[int, slice]) -> Optional["Character"]:
        """Getting the character at the given index, including the delimiter characters."""

        if isinstance(index, slice):
            start = index.start if index.start is not None else 0
            stop = index.stop if index.stop is not None else len(self)
            step = index.step if index.step is not None else 1

            if start < 0:
                start = len(self) + start
            if stop < 0:
                stop = len(self) + stop
            return [self.__getitem__(i) for i in range(start, stop, step)]
        if isinstance(index, int):
            if index < 0 or index >= len(self):
                raise IndexError("Index out of range of text length")
            return self._get_char(index)
        raise TypeError("Index must be an integer or a slice")

    def _get_char(self, index: int) -> Optional["Character"]:
        for child in self.children:
            if 0 <= index < child.length:
                return child._get_char(index)
            index -= child.length + len(child.delimiter)
        return None

    def set_absolute_order(self, absolute_order: int):
        """
        Set the absolute order of the block in the document.
        """
        self.absolute_order = absolute_order

    def get_absolute_order(self) -> int:
        """
        Returns the absolute order of the block in the document.
        The absolute order of a text block is the number of blocks
        of the same type that precede it in the document.
        """
        return self.absolute_order

    def set_start_index(self) -> int:
        """
        Set the index of the first character of the block in the document.
        """
        if self.parent is None:
            self.start_index = 0
            return
        if self.relative_order == 0:
            self.start_index = self.parent.get_start_index()
            return
        previous_child = self.parent.children[self.relative_order - 1]
        self.start_index = previous_child.get_end_index() + len(self.delimiter)

    def get_start_index(self):
        """
        Returns the index of the first character of the block in the document.
        """
        if self.start_index is None:
            self.set_start_index()
        return self.start_index

    def set_end_index(self):
        """
        Set the index of the last character of the block in the document.
        """
        self.end_index = self.get_start_index() + self.get_length()

    def get_end_index(self) -> int:
        if self.end_index is None:
            self.set_end_index()
        return self.end_index


class Character(TextBlock):

    """
    The Character class represents a single character (not including delimiters).
    The Character class is the leaf of the TextBlock tree.
    """

    def __init__(self, bounding_box: BoundingBox, char: str):
        super().__init__(bounding_box=bounding_box)
        self.char = char
        self.verify_args()
        self.length = 1

    def verify_args(self):
        if not isinstance(self.char, str):
            raise ValueError("char must be a string")
        if len(self.char) != 1:
            logging.warning(
                "char %s is not a single character. It's length is %s",
                self.char,
                len(self.char),
            )

    @staticmethod
    def get_char_weight(char: str) -> float:
        if char.islower():
            return CHARACTER_WIDTH_MAP["lower"]
        if char.isupper():
            return CHARACTER_WIDTH_MAP["upper"]
        if char in THIN_CHARS:
            return CHARACTER_WIDTH_MAP["thin"]
        return CHARACTER_WIDTH_MAP["default"]

    def get_text(self) -> str:
        return self.char

    def set_length(self) -> int:
        self.length = 1

    def _get_char(self, index: int) -> str:
        return self

    def __len__(self):
        return 1


class Word(TextBlock):
    """
    The Word class represents a single word.
    """

    child_type = Character
    delimiter = WORD_DELIMITER

    def __init__(
        self,
        text: str,
        bounding_box: BoundingBox,
        printed: bool = True,
    ):
        super().__init__(bounding_box=bounding_box)
        text = self.get_clean_text(text)
        char_weights = np.array(
            [Character.get_char_weight(char) for char in text]
        )
        chars_bboxes = BoundingBox.split_bounding_box(
            bounding_box, len(text), weights=char_weights
        )
        self.text = text
        self.printed = printed
        for char, char_bbox in zip(text, chars_bboxes):
            self.set_child(Character(char_bbox, char))
        self.length = len(self.text)

    def get_clean_text(self, text: str) -> str:
        clean_text = [LIGATURES_TO_TEXT.get(char, char) for char in text]
        clean_text = "".join(clean_text)
        return clean_text

    def __len__(self):
        return len(self.text)


class Line(TextBlock):
    """
    The Line class represents a single line.
    """

    child_type = Word
    delimiter = LINE_DELIMITER


class Area(TextBlock, metaclass=abc.ABCMeta):
    delimiter = AREA_DELIMITER
    """
    The Area class represents a single area.
    """


class Paragraph(Area):
    """
    The Paragraph class represents a single paragraph.
    """

    child_type = Line
    delimiter = PARAGRAPH_DELIMITER


class Page(TextBlock):
    """
    The Page class represents a single page.
    """

    child_type = Area
    delimiter = PAGE_DELIMITER

    def __init__(
        self, children: list[Area], width: float = 1, height: float = 1
    ):
        super().__init__(
            bounding_box=BoundingBox(**NORMALIZED_BBOX),
            children=children,
        )
        self.width = width
        self.height = height

    def set_bounding_box(self):
        """The bounding box of the page is fixed to BoundingBox(left=0, top=1, right=1, bottom=0).
        The coordinates of the blocks are PDF coordinates (i.e. the origin is at the bottom left),
        and normalized to the page size."""
        return

    def normalize_point(self, point: Point) -> Point:
        """
        Normalize a point from PDF coordinates to normalized coordinates.
        """
        return Point(x=point.x / self.width, y=point.y / self.height)

    def denormalize_point(self, point: Point) -> Point:
        """
        Denormalize a point from normalized coordinates to PDF coordinates.
        """
        return Point(x=point.x * self.width, y=point.y * self.height)


class Document(TextBlock):
    """
    The Document class represents a single document.
    """

    child_type = Page

    def __init__(self, children: list[Page]):
        super().__init__(children=children)

        self.set_length()
        for text_block in TextBlock.pre_order_traversal(self):
            text_block.set_start_index()
            text_block.set_end_index()

        block_types = [Page, Area, Line, Word, Character]
        for block_type in block_types:
            text_blocks = self.get_all(block_type)
            for i, text_block in enumerate(text_blocks):
                text_block.set_absolute_order(i)

        self.text = self.get_text()

    def get_intersecting_text_blocks(
        self, _class: type, bbox: BoundingBox, page_number: int
    ) -> Generator:
        if page_number < 0 or page_number >= len(self.children):
            raise ValueError(
                "page_number must be between 0 and the number of pages"
            )
        page: Page = self.children[page_number]
        for text_block in page.get_intersecting_text_blocks(_class, bbox):
            yield text_block

    def set_bounding_box(self):
        """There is no bounding box for the document, therefore the method does nothing."""
        return

    def get_text_block(
        self, index: int, text_block_type: type
    ) -> Optional[TextBlock]:
        """
        Get the text block of type `text_block_type` that contains the character at index `index`.
        The method handle the case where the character is a delimiter as well.
        """
        character = self[index]
        if character is not None:
            text_block = character.get_ancestor(text_block_type)
            return text_block
        previous_char = self[index - 1]
        next_char = self[index + 1]
        if previous_char is None or next_char is None:
            return None
        previous_text_block = previous_char.get_ancestor(text_block_type)
        next_text_block = next_char.get_ancestor(text_block_type)
        if previous_text_block == next_text_block:
            return previous_text_block
        return None

    def get_word(self, index: int) -> Optional[Word]:
        """
        Get the word object that contains the character at index `index`.
        In case the character is a word delimiter, the return value is None.
        """
        return self.get_text_block(index, Word)

    def get_line(self, index: int) -> Optional[Line]:
        """
        Get the line object that contains the character at index `index`,
        including the case where the character is a word delimiter (space).
        In case the character is a line delimiter, the return value is None.
        """
        return self.get_text_block(index, Line)

    def get_paragraph(self, index: int) -> Optional[Paragraph]:
        """
        Get the paragraph object that contains the character at index `index`,
        including the case where the character is a word delimiter (space)
        or a line delimiter (newline).
        In case the character is a paragraph delimiter, the return value is None.

        """
        return self.get_text_block(index, Paragraph)

    def get_page(self, index: int) -> Optional[Page]:
        """
        Get the page object that contains the character at index `index`,
        including the case where the character is a word delimiter (space),
        a line delimiter (newline) or a paragraph delimiter.
        In case the character is a page delimiter, the return value is None.

        """
        return self.get_text_block(index, Page)


class TableCell(TextBlock):
    child_type = Line
    delimiter = CELL_DELIMITER

    def __init__(
        self,
        bounding_box: BoundingBox,
        children: list[Line],
        row_index: int,
        col_index: int,
        row_span: int,
        col_span: int,
    ):
        super().__init__(bounding_box=bounding_box, children=children)
        self.row_index = row_index
        self.col_index = col_index
        self.row_span = row_span
        self.col_span = col_span


class TableColumn(TextBlock):
    child_type = TableCell
    delimiter = COLUMN_DELIMITER


class TableType(Enum):
    BORDERED_TABLE = "bordered_table"
    BORDERLESS_TABLE = "borderless_table"

    def __str__(self):
        return self.value


class Table(Area):
    child_type = TableColumn
    delimiter = TABLE_DELIMITER

    def __init__(
        self,
        children: list[TableColumn],
        table_type: TableType = TableType.BORDERED_TABLE,
    ):
        super().__init__(children=children)
        self.verify_args(children)
        self.grid = self.get_grid()
        self.table_type = table_type

    def verify_args(self, children: list[TableColumn]):
        lines_set = set()
        for column in children:
            column_lines = column.get_all(Line)
            for line in column_lines:
                if line in lines_set:
                    raise ValueError("Duplicate line in a table")
                lines_set.add(line)

    def get_grid(self) -> np.ndarray:
        cells: list[TableCell] = list(self.get_all(TableCell))
        num_rows = (
            max(cell.row_index + cell.row_span - 1 for cell in cells) + 1
        )
        num_columns = (
            max(cell.col_index + cell.col_span - 1 for cell in cells) + 1
        )
        grid = np.empty((num_rows, num_columns), dtype=object)
        for cell in cells:
            for row in range(cell.row_index, cell.row_index + cell.row_span):
                for col in range(
                    cell.col_index, cell.col_index + cell.col_span
                ):
                    grid[row, col] = cell
        return grid

    def to_csv(self, filepath: str):
        cells: Generator[TableCell, None, None] = self.get_all(TableCell)
        max_row_index = max(
            cell.row_index + cell.row_span - 1 for cell in cells
        )
        max_col_index = max(
            cell.col_index + cell.col_span - 1 for cell in cells
        )
        table_array = [
            ["" for _ in range(max_col_index + 1)]
            for _ in range(max_row_index + 1)
        ]

        for cell in cells:
            cell_str = re.sub(r"\s", " ", str(cell))
            table_array[cell.row_index][cell.col_index] = cell_str

        with open(filepath, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            for row in table_array:
                writer.writerow(row)

    def to_xlsx_obj(self) -> Workbook:
        work_book = Workbook()
        ws = work_book.active
        cells: Generator[TableCell, None, None] = self.get_all(TableCell)
        # Write the table to the worksheet
        for cell in cells:
            cell_str = str(cell)
            # Write the cell value
            col_letter = get_column_letter(cell.col_index + 1)
            cell_ref = col_letter + str(cell.row_index + 1)
            ws[cell_ref] = cell_str

            # Apply cell formatting
            ws[cell_ref].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            if cell.row_span > 1 or cell.col_span > 1:
                # Merge cells if necessary
                end_col_letter = get_column_letter(
                    cell.col_index + cell.col_span
                )
                end_row_index = cell.row_index + cell.row_span - 1
                end_cell_ref = end_col_letter + str(end_row_index + 1)
                ws.merge_cells(cell_ref + ":" + end_cell_ref)
        return work_book

    def to_xlsx(self, filepath: str):
        work_book = self.to_xlsx_obj()
        work_book.save(filepath)
