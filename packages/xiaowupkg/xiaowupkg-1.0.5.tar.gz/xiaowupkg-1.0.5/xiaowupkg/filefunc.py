import datetime
import os
from openpyxl import load_workbook

def readfilename(path, filetype):     #获取ktr文件
    L=[]
    for root,dirs,files in os.walk(path):
        if root == path:
            for i in files:
                if os.path.splitext(i)[1] == filetype:
                    print(i)
                    L.append(i)
    return L

'''
功能：清除当前路径下所有的python缓存文件
'''
def cleanPycache(path):
    for root,dirs,file in os.walk(path):
        print(root)
        print(dirs)
        for dir in dirs:
            if dir == '__pycache__':
                os.system('rm -rf ' + root + os.sep + dir)
                print('clear finished:' + root + os.sep + dir)
        print('\n')

'''
功能：读取指定路径xlsx文件
'''
def readxlsx(path):
    global clients
    # 读取Excel文件
    workbook = load_workbook(path)
    worksheet = workbook['Sheet1']
    
    # 打印数据
    for cols in worksheet.iter_cols(values_only=True):
        coltext = []
        for col in cols:
            if type(col) is not datetime.time and col is not None:
                coltext.append(col)
        clients.append(coltext)
    print(clients)
